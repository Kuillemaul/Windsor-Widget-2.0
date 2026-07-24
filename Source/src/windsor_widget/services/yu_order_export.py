"""Validated supplier-facing Yuchang order-form export.

The YU workbook's Sheet1 column A is the permanent Windsor item mapping. Row
numbers are treated only as current workbook addresses and are resolved again
immediately before every export.
"""

from __future__ import annotations

import csv
import json
import math
import os
import re
import shutil
import tempfile
import uuid
import zipfile
from collections import defaultdict
from copy import copy, deepcopy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from windsor_widget.db.models import (
    AppUser,
    AuditEvent,
    ItemSupplier,
    ManufactureOrder,
    ManufactureOrderLine,
)
from windsor_widget.db.models.audit import utc_now

SHEET_NAME = "Sheet1"
HEADER_END_ROW = 14
ITEM_COLUMN = "A"
QUANTITY_COLUMN = "L"
DATE_CELL = "C10"
ORDER_NUMBER_CELL = "H10"
EXPORT_MIN_COLUMN = "A"
EXPORT_MAX_COLUMN = "N"


class YUWorkbookChanged(RuntimeError):
    """Raised when the master workbook changed after the validation screen loaded."""


@dataclass(frozen=True, slots=True)
class YUWorkbookRow:
    row_number: int
    mapped_item_number: str
    item: str
    size: str
    colour: str
    pack_type: str
    unit_quantity: str
    labelled_as: str
    metres_per_carton: str
    metres_per_pallet: str
    pallet_count: str
    unit_price: str

    @property
    def description(self) -> str:
        return " · ".join(
            value
            for value in (
                self.item,
                self.size,
                self.colour,
                self.pack_type,
                self.unit_quantity,
                self.labelled_as,
            )
            if value
        )


@dataclass(frozen=True, slots=True)
class YUCandidate:
    row: YUWorkbookRow
    source: str
    score: float | None


@dataclass(frozen=True, slots=True)
class YUOrderLineValidation:
    line_id: uuid.UUID
    item_id: uuid.UUID
    item_number: str
    item_name: str
    item_description: str
    quantity: Decimal
    status: str
    status_label: str
    mapped_rows: tuple[YUWorkbookRow, ...]
    candidates: tuple[YUCandidate, ...]


@dataclass(frozen=True, slots=True)
class YUOrderValidationReport:
    order_id: uuid.UUID
    order_number: str
    supplier_name: str
    template_path: str
    worksheet_name: str
    workbook_mtime_ns: int
    workbook_modified_at: datetime
    header_end_row: int
    footer_start_row: int
    footer_end_row: int
    lines: tuple[YUOrderLineValidation, ...]
    resolved_count: int
    missing_count: int
    duplicate_count: int
    ready: bool


@dataclass(frozen=True, slots=True)
class YUExportResult:
    output_path: str
    audit_path: str
    item_count: int
    line_count: int
    workbook_mtime_ns: int
    resolved_rows: tuple[tuple[str, int], ...]


def clean_item_key(value: str) -> str:
    """Space-insensitive key retained from the proven Widget 1 YU workflow."""
    return re.sub(r"[\s\u00A0]+", "", str(value or "").strip()).upper()


def _normalise_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value):
            return format(value, ".12g").upper()
        return str(value).upper()
    return re.sub(r"[\s\u00A0]+", " ", str(value).strip()).upper()


def _display(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_is_detail(values: tuple[Any, ...] | list[Any]) -> bool:
    if not values or values[0] in (None, ""):
        return False
    return any(value not in (None, "") for value in values[1:7])


def _footer_rows_from_cache(
    row_cache: list[tuple[int, tuple[Any, ...]]],
    *,
    header_end_row: int,
) -> tuple[int, int]:
    transport_row: int | None = None
    total_row: int | None = None
    for row_number, values in row_cache:
        normalized = [_normalise_text(value) for value in values[:14]]
        if total_row is None and any(
            value == "ORDER TOTAL" or value.startswith("ORDER TOTAL ")
            for value in normalized
        ):
            total_row = row_number
        if transport_row is None and any(
            "METHOD OF TRANSPORT" in value for value in normalized
        ):
            transport_row = row_number
    last_row = row_cache[-1][0] if row_cache else header_end_row
    if total_row is None:
        total_row = last_row
    if transport_row is None or transport_row > total_row:
        transport_row = max(header_end_row + 1, total_row - 5)
    footer_start = transport_row
    previous_values = next(
        (values for row_number, values in row_cache if row_number == footer_start - 1),
        (),
    )
    if previous_values and all(value in (None, "") for value in previous_values[:14]):
        footer_start -= 1
    return int(footer_start), int(total_row)


def scan_yu_workbook(
    template_path: str | Path,
    *,
    worksheet_name: str = SHEET_NAME,
) -> dict[str, Any]:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"YU workbook was not found: {path}")
    keep_vba = path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=keep_vba,
    )
    try:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {worksheet_name!r} was not found. Available sheets: "
                + ", ".join(workbook.sheetnames)
            )
        sheet = workbook[worksheet_name]
        row_cache: list[tuple[int, tuple[Any, ...]]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(
                min_row=HEADER_END_ROW + 1,
                max_row=int(sheet.max_row or 0),
                min_col=1,
                max_col=14,
                values_only=True,
            ),
            start=HEADER_END_ROW + 1,
        ):
            row_cache.append((row_number, tuple(values)))
        footer_start, footer_end = _footer_rows_from_cache(
            row_cache,
            header_end_row=HEADER_END_ROW,
        )

        exact_rows: dict[str, list[int]] = defaultdict(list)
        clean_rows: dict[str, list[int]] = defaultdict(list)
        rows: dict[int, YUWorkbookRow] = {}
        for row_number, values in row_cache:
            if row_number >= footer_start:
                break
            detail_values = list(values[1:8])
            if not _row_is_detail(detail_values):
                continue
            mapped = _display(values[0])
            row = YUWorkbookRow(
                row_number=row_number,
                mapped_item_number=mapped,
                item=_display(values[1]),
                size=_display(values[2]),
                colour=_display(values[3]),
                pack_type=_display(values[4]),
                unit_quantity=_display(values[5]),
                labelled_as=_display(values[6]),
                metres_per_carton=_display(values[7]),
                metres_per_pallet=_display(values[8]),
                pallet_count=_display(values[9]),
                unit_price=_display(values[10]),
            )
            rows[row_number] = row
            if mapped:
                exact_rows[mapped.upper()].append(row_number)
                clean_rows[clean_item_key(mapped)].append(row_number)

        review_candidates: dict[str, list[tuple[int, str, float | None]]] = defaultdict(list)
        if "Match_Review" in workbook.sheetnames:
            review = workbook["Match_Review"]
            for values in review.iter_rows(min_row=5, values_only=True):
                if not values:
                    continue
                try:
                    source_row = int(float(values[0]))
                except (TypeError, ValueError):
                    continue
                final = str(values[1] or "").strip() if len(values) > 1 else ""
                suggested = str(values[2] or "").strip() if len(values) > 2 else ""
                confidence = None
                if len(values) > 3 and values[3] not in (None, ""):
                    try:
                        confidence = float(values[3])
                    except (TypeError, ValueError):
                        confidence = None
                candidates = [
                    (final, "Match_Review final selection"),
                    (suggested, "Match_Review suggested match"),
                ]
                for index in (13, 15, 17, 19, 21):
                    if len(values) > index:
                        candidates.append(
                            (
                                str(values[index] or "").strip(),
                                "Match_Review candidate",
                            )
                        )
                for item_number, source in candidates:
                    key = clean_item_key(item_number)
                    if key:
                        review_candidates[key].append((source_row, source, confidence))

        stat = path.stat()
        return {
            "path": str(path),
            "worksheet_name": worksheet_name,
            "sheet_names": tuple(workbook.sheetnames),
            "mtime_ns": int(stat.st_mtime_ns),
            "modified_at": datetime.fromtimestamp(stat.st_mtime),
            "header_end_row": HEADER_END_ROW,
            "footer_start_row": footer_start,
            "footer_end_row": footer_end,
            "exact_rows": dict(exact_rows),
            "clean_rows": dict(clean_rows),
            "rows": rows,
            "review_candidates": dict(review_candidates),
        }
    finally:
        workbook.close()


def _rows_for_item(scan: dict[str, Any], item_number: str) -> list[int]:
    item = str(item_number or "").strip()
    if not item:
        return []
    exact = list((scan.get("exact_rows") or {}).get(item.upper(), []))
    if exact:
        return sorted(set(int(row) for row in exact))
    return sorted(
        set(
            int(row)
            for row in (scan.get("clean_rows") or {}).get(clean_item_key(item), [])
        )
    )


def _match_review_candidates(
    template_path: str | Path,
    item_number: str,
) -> list[tuple[int, str, float | None]]:
    path = Path(template_path)
    keep_vba = path.suffix.casefold() == ".xlsm"
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_vba=keep_vba,
    )
    try:
        if "Match_Review" not in workbook.sheetnames:
            return []
        sheet = workbook["Match_Review"]
        wanted = clean_item_key(item_number)
        results: list[tuple[int, str, float | None]] = []
        for values in sheet.iter_rows(min_row=5, values_only=True):
            if not values:
                continue
            try:
                source_row = int(float(values[0]))
            except (TypeError, ValueError):
                continue
            final = str(values[1] or "").strip() if len(values) > 1 else ""
            suggested = str(values[2] or "").strip() if len(values) > 2 else ""
            confidence = None
            if len(values) > 3 and values[3] not in (None, ""):
                try:
                    confidence = float(values[3])
                except (TypeError, ValueError):
                    confidence = None
            candidate_values = [final, suggested]
            for index in (13, 15, 17, 19, 21):
                if len(values) > index:
                    candidate_values.append(str(values[index] or "").strip())
            matched = [value for value in candidate_values if clean_item_key(value) == wanted]
            if not matched:
                continue
            source = "Match_Review final selection" if clean_item_key(final) == wanted else (
                "Match_Review suggested match"
                if clean_item_key(suggested) == wanted
                else "Match_Review candidate"
            )
            results.append((source_row, source, confidence))
        return results
    finally:
        workbook.close()


def _tokens(value: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[A-Z0-9]+", _normalise_text(value))
        if len(token) > 1
    }


def _candidate_score(target_text: str, row: YUWorkbookRow) -> float:
    row_text = row.description
    ratio = SequenceMatcher(None, _normalise_text(target_text), _normalise_text(row_text)).ratio()
    target_tokens = _tokens(target_text)
    row_tokens = _tokens(row_text)
    overlap = (
        len(target_tokens & row_tokens) / len(target_tokens | row_tokens)
        if target_tokens and row_tokens
        else 0.0
    )
    return max(0.0, min(1.0, ratio * 0.55 + overlap * 0.45))


def _candidate_rows(
    scan: dict[str, Any],
    *,
    template_path: str | Path,
    item_number: str,
    item_name: str,
    item_description: str,
    supplier_item_number: str,
    limit: int = 10,
) -> tuple[YUCandidate, ...]:
    rows: dict[int, YUWorkbookRow] = scan["rows"]
    selected: dict[int, YUCandidate] = {}
    for row_number, source, confidence in (
        (scan.get("review_candidates") or {}).get(clean_item_key(item_number), [])
    ):
        row = rows.get(row_number)
        if row is not None:
            selected[row_number] = YUCandidate(
                row=row,
                source=source,
                score=confidence,
            )

    target = " ".join(
        value
        for value in (item_number, item_name, item_description, supplier_item_number)
        if value
    )
    scored = sorted(
        (
            (_candidate_score(target, row), row)
            for row in rows.values()
            if row.row_number not in selected
        ),
        key=lambda value: (value[0], -value[1].row_number),
        reverse=True,
    )
    for score, row in scored:
        if len(selected) >= limit:
            break
        if score < 0.12 and selected:
            break
        selected[row.row_number] = YUCandidate(
            row=row,
            source="Description similarity",
            score=score,
        )
    return tuple(
        sorted(
            selected.values(),
            key=lambda value: (
                value.score if value.score is not None else -1,
                -value.row.row_number,
            ),
            reverse=True,
        )[:limit]
    )


def _load_order(session: Session, order_id: uuid.UUID) -> ManufactureOrder:
    order = session.scalar(
        select(ManufactureOrder)
        .options(
            selectinload(ManufactureOrder.lines).selectinload(
                ManufactureOrderLine.item
            )
        )
        .where(ManufactureOrder.manufacture_order_id == order_id)
    )
    if order is None:
        raise LookupError("Manufacture order not found.")
    return order


def validate_yu_order(
    session: Session,
    *,
    order_id: uuid.UUID,
    template_path: str | Path,
    worksheet_name: str = SHEET_NAME,
) -> YUOrderValidationReport:
    order = _load_order(session, order_id)
    scan = scan_yu_workbook(template_path, worksheet_name=worksheet_name)
    rows: dict[int, YUWorkbookRow] = scan["rows"]
    validations: list[YUOrderLineValidation] = []

    for line in order.lines:
        quantity = Decimal(line.ordered_quantity) - Decimal(line.cancelled_quantity)
        if quantity <= 0:
            continue
        mapped_numbers = _rows_for_item(scan, line.item.item_number)
        mapped_rows = tuple(rows[row_number] for row_number in mapped_numbers if row_number in rows)
        if len(mapped_rows) == 1:
            status = "resolved"
            status_label = f"Resolved to row {mapped_rows[0].row_number}"
            candidates: tuple[YUCandidate, ...] = ()
        elif len(mapped_rows) > 1:
            status = "duplicate"
            status_label = "Duplicate Column A mapping"
            candidates = tuple(
                YUCandidate(row=row, source="Current Column A mapping", score=1.0)
                for row in mapped_rows
            )
        else:
            status = "missing"
            status_label = "Mapping required"
            supplier_item_number = session.scalar(
                select(ItemSupplier.supplier_item_number).where(
                    ItemSupplier.supplier_id == order.supplier_id,
                    ItemSupplier.item_id == line.item_id,
                    ItemSupplier.match_status != "rejected",
                )
            ) or ""
            candidates = _candidate_rows(
                scan,
                template_path=template_path,
                item_number=line.item.item_number,
                item_name=line.item.item_name,
                item_description=line.item.description or "",
                supplier_item_number=supplier_item_number,
            )
        validations.append(
            YUOrderLineValidation(
                line_id=line.manufacture_order_line_id,
                item_id=line.item_id,
                item_number=line.item.item_number,
                item_name=line.item.item_name,
                item_description=line.item.description or "",
                quantity=quantity,
                status=status,
                status_label=status_label,
                mapped_rows=mapped_rows,
                candidates=candidates,
            )
        )

    resolved = sum(line.status == "resolved" for line in validations)
    missing = sum(line.status == "missing" for line in validations)
    duplicates = sum(line.status == "duplicate" for line in validations)
    return YUOrderValidationReport(
        order_id=order.manufacture_order_id,
        order_number=order.order_number,
        supplier_name=order.supplier.display_name,
        template_path=str(template_path),
        worksheet_name=worksheet_name,
        workbook_mtime_ns=int(scan["mtime_ns"]),
        workbook_modified_at=scan["modified_at"],
        header_end_row=int(scan["header_end_row"]),
        footer_start_row=int(scan["footer_start_row"]),
        footer_end_row=int(scan["footer_end_row"]),
        lines=tuple(validations),
        resolved_count=resolved,
        missing_count=missing,
        duplicate_count=duplicates,
        ready=bool(validations) and missing == 0 and duplicates == 0,
    )


# ---------------------------------------------------------------------------
# Safe direct-XML mapping update. This changes only Sheet1 column A and avoids
# an openpyxl round-trip of the complex multi-sheet master workbook.
# ---------------------------------------------------------------------------
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
XLSX_IGNORABLE_NAMESPACES = {
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
}
ET_NS = {"s": SHEET_NS, "r": OFFICE_REL_NS, "rel": PACKAGE_REL_NS}


def _register_xlsx_namespaces() -> None:
    import xml.etree.ElementTree as ET

    ET.register_namespace("", SHEET_NS)
    ET.register_namespace("r", OFFICE_REL_NS)
    ET.register_namespace("mc", MC_NS)
    for prefix, uri in XLSX_IGNORABLE_NAMESPACES.items():
        ET.register_namespace(prefix, uri)


def _namespace_used(root: Any, uri: str) -> bool:
    marker = f"{{{uri}}}"
    for node in root.iter():
        if str(node.tag).startswith(marker):
            return True
        if any(str(name).startswith(marker) for name in node.attrib):
            return True
    return False


def _xlsx_xml_bytes(root: Any) -> bytes:
    import xml.etree.ElementTree as ET

    ignorable_name = f"{{{MC_NS}}}Ignorable"
    for prefix in str(root.attrib.get(ignorable_name, "") or "").split():
        uri = XLSX_IGNORABLE_NAMESPACES.get(prefix)
        if uri and not _namespace_used(root, uri):
            root.set(f"xmlns:{prefix}", uri)
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return data.replace(
        b"<?xml version='1.0' encoding='utf-8'?>",
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        1,
    )


def _sheet_part(zip_file: zipfile.ZipFile, sheet_name: str) -> str:
    import xml.etree.ElementTree as ET

    workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
    rels = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    targets = {rel.attrib.get("Id"): rel.attrib.get("Target", "") for rel in rels}
    sheets = workbook.find("s:sheets", ET_NS)
    if sheets is None:
        raise ValueError("Workbook contains no sheet list.")
    for sheet in sheets.findall("s:sheet", ET_NS):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{OFFICE_REL_NS}}}id")
        target = targets.get(rel_id, "")
        if target.startswith("/"):
            return target.lstrip("/")
        if target:
            return str(PurePosixPath("xl") / target)
    raise ValueError(f"Sheet {sheet_name!r} not found in workbook.")


def _split_cell_ref(cell_ref: str) -> tuple[str, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", str(cell_ref or "").upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    return match.group(1), int(match.group(2))


def _cell_sort_key(cell: Any) -> int:
    return column_index_from_string(_split_cell_ref(cell.attrib.get("r", "A1"))[0])


def _rows_by_number(root: Any) -> dict[int, Any]:
    data = root.find("s:sheetData", ET_NS)
    if data is None:
        raise ValueError("Worksheet has no sheetData node.")
    rows: dict[int, Any] = {}
    for row in data.findall("s:row", ET_NS):
        try:
            rows[int(row.attrib.get("r", "0"))] = row
        except ValueError:
            continue
    return rows


def _ensure_xml_row(root: Any, row_number: int) -> Any:
    import xml.etree.ElementTree as ET

    data = root.find("s:sheetData", ET_NS)
    if data is None:
        data = ET.SubElement(root, f"{{{SHEET_NS}}}sheetData")
    existing = _rows_by_number(root).get(row_number)
    if existing is not None:
        return existing
    row = ET.Element(f"{{{SHEET_NS}}}row", {"r": str(row_number)})
    for index, current in enumerate(list(data)):
        try:
            if int(current.attrib.get("r", "0")) > row_number:
                data.insert(index, row)
                return row
        except ValueError:
            continue
    data.append(row)
    return row


def _ensure_xml_cell(root: Any, cell_ref: str) -> Any:
    import xml.etree.ElementTree as ET

    _, row_number = _split_cell_ref(cell_ref)
    row = _ensure_xml_row(root, row_number)
    for cell in row.findall("s:c", ET_NS):
        if cell.attrib.get("r") == cell_ref:
            return cell
    cell = ET.Element(f"{{{SHEET_NS}}}c", {"r": cell_ref})
    target_key = _cell_sort_key(cell)
    for index, current in enumerate(list(row)):
        if _cell_sort_key(current) > target_key:
            row.insert(index, cell)
            return cell
    row.append(cell)
    return cell


def _set_xml_cell_text(root: Any, cell_ref: str, value: str | None) -> None:
    import xml.etree.ElementTree as ET

    cell = _ensure_xml_cell(root, cell_ref)
    for child in list(cell):
        if child.tag in {
            f"{{{SHEET_NS}}}v",
            f"{{{SHEET_NS}}}is",
            f"{{{SHEET_NS}}}f",
        }:
            cell.remove(child)
    text = str(value or "").strip()
    if not text:
        cell.attrib.pop("t", None)
        return
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, f"{{{SHEET_NS}}}is")
    node = ET.SubElement(inline, f"{{{SHEET_NS}}}t")
    node.text = text


def _write_mapping_cells(
    workbook_path: Path,
    row_values: dict[int, str | None],
    *,
    worksheet_name: str,
) -> None:
    import xml.etree.ElementTree as ET

    _register_xlsx_namespaces()
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}_",
        suffix=workbook_path.suffix,
        dir=str(workbook_path.parent),
    )
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with zipfile.ZipFile(workbook_path, "r") as source:
            part = _sheet_part(source, worksheet_name)
            root = ET.fromstring(source.read(part))
            for row_number, item_number in row_values.items():
                _set_xml_cell_text(root, f"A{int(row_number)}", item_number)
            replacement = _xlsx_xml_bytes(root)
            with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
                for info in source.infolist():
                    data = replacement if info.filename == part else source.read(info.filename)
                    target.writestr(info, data)
        os.replace(temp_path, workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _actor(session: Session, actor_user_id: uuid.UUID) -> AppUser:
    actor = session.get(AppUser, actor_user_id)
    if actor is None or not actor.is_active:
        raise LookupError("The signed-in user is no longer active.")
    return actor


def apply_yu_item_mapping(
    session: Session,
    *,
    order_id: uuid.UUID,
    line_id: uuid.UUID,
    source_row: int,
    template_path: str | Path,
    expected_mtime_ns: int,
    actor_user_id: uuid.UUID,
    worksheet_name: str = SHEET_NAME,
    clear_other_item_rows: bool = False,
) -> tuple[str, tuple[int, ...], str]:
    actor = _actor(session, actor_user_id)
    order = _load_order(session, order_id)
    line = next(
        (value for value in order.lines if value.manufacture_order_line_id == line_id),
        None,
    )
    if line is None:
        raise LookupError("Manufacture-order line not found.")
    path = Path(template_path)
    if path.stat().st_mtime_ns != int(expected_mtime_ns):
        raise YUWorkbookChanged(
            "The YU workbook changed after this page was loaded. "
            "Revalidate before saving a mapping."
        )
    scan = scan_yu_workbook(path, worksheet_name=worksheet_name)
    candidate: YUWorkbookRow | None = scan["rows"].get(int(source_row))
    if candidate is None:
        raise ValueError(
            f"Row {source_row} is not a current YU product/detail row in {worksheet_name}."
        )
    item_number = line.item.item_number
    occupied = candidate.mapped_item_number
    if occupied and clean_item_key(occupied) != clean_item_key(item_number):
        raise ValueError(
            f"Row {source_row} is already mapped to {occupied}. Choose another row."
        )
    current_rows = _rows_for_item(scan, item_number)
    changes: dict[int, str | None] = {int(source_row): item_number}
    cleared: list[int] = []
    if clear_other_item_rows:
        for row_number in current_rows:
            if row_number != int(source_row):
                changes[row_number] = None
                cleared.append(row_number)
    elif current_rows and int(source_row) not in current_rows:
        raise ValueError(
            f"{item_number} is already mapped to row(s) "
            + ", ".join(str(row) for row in current_rows)
            + ". Use the duplicate-resolution action to move the mapping."
        )

    backup_dir = path.parent / "_Widget Backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / (
        f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S_%f}{path.suffix}"
    )
    shutil.copy2(path, backup_path)
    try:
        _write_mapping_cells(path, changes, worksheet_name=worksheet_name)
        rescanned = scan_yu_workbook(path, worksheet_name=worksheet_name)
        resolved = _rows_for_item(rescanned, item_number)
        if resolved != [int(source_row)]:
            raise RuntimeError(
                "The workbook was saved but the mapping did not revalidate uniquely. "
                f"Current rows: {resolved}"
            )
    except Exception:
        shutil.copy2(backup_path, path)
        raise

    session.add(
        AuditEvent(
            actor_user_id=actor.user_id,
            action="supplier_order_template.mapping_updated",
            entity_type="manufacture_order_line",
            entity_id=str(line.manufacture_order_line_id),
            source="web",
            summary=(
                f"Mapped {item_number} to YU {worksheet_name} row {source_row}."
                + (
                    " Cleared duplicate row(s) " + ", ".join(str(row) for row in cleared) + "."
                    if cleared
                    else ""
                )
            )[:500],
            before_json=json.dumps(
                {
                    "template_path": str(path),
                    "mapped_rows": current_rows,
                    "selected_row_previous_value": occupied or None,
                },
                sort_keys=True,
            ),
            after_json=json.dumps(
                {
                    "template_path": str(path),
                    "mapped_row": int(source_row),
                    "cleared_rows": cleared,
                    "backup_path": str(backup_path),
                },
                sort_keys=True,
            ),
        )
    )
    session.flush()
    return item_number, tuple(cleared), str(backup_path)


# ---------------------------------------------------------------------------
# Compact supplier workbook creation.
# ---------------------------------------------------------------------------
def _collect_header_note_rows(
    source_row: int,
    *,
    header_end_row: int,
    matched_rows: set[int],
    row_has_item,
    row_has_content,
    max_blank_scan: int = 3,
    max_unmatched_item_scan: int = 25,
) -> set[int]:
    related: set[int] = set()
    scan = source_row - 1
    blanks = 0
    skipped_items = 0
    found_note = False
    while scan > header_end_row:
        if row_has_item(scan):
            if scan in matched_rows:
                break
            skipped_items += 1
            if skipped_items > max_unmatched_item_scan:
                break
            blanks = 0
            scan -= 1
            continue
        if row_has_content(scan):
            related.add(scan)
            found_note = True
            blanks = 0
            scan -= 1
            continue
        blanks += 1
        if found_note or skipped_items > 0 or blanks > max_blank_scan:
            break
        scan -= 1
    return related


def _row_plan(
    kept_rows: list[int],
    row_has_item,
    *,
    header_end_row: int,
    footer_start_row: int,
    separator_source_gap: int = 50,
) -> list[int | None]:
    result: list[int | None] = []
    previous_detail: int | None = None
    for old_row in kept_rows:
        detail = header_end_row < old_row < footer_start_row and row_has_item(old_row)
        if (
            detail
            and previous_detail is not None
            and old_row - previous_detail > separator_source_gap
        ):
            previous = result[-1] if result else None
            previous_is_note = (
                isinstance(previous, int)
                and header_end_row < previous < footer_start_row
                and not row_has_item(previous)
            )
            if not previous_is_note:
                result.append(None)
        result.append(old_row)
        if detail:
            previous_detail = old_row
    return result


def _cell_text_for_width(cell: Any) -> str:
    value = cell.value
    if value in (None, ""):
        return ""
    if isinstance(value, str) and value.startswith("="):
        match = re.fullmatch(
            r"=\s*([A-Z]+)(\d+)\s*\*\s*([A-Z]+)(\d+)\s*",
            value.strip(),
            flags=re.IGNORECASE,
        )
        if not match:
            return ""
        try:
            left = cell.parent[f"{match.group(1).upper()}{match.group(2)}"].value
            right = cell.parent[f"{match.group(3).upper()}{match.group(4)}"].value
            if isinstance(left, (int, float)) and isinstance(right, (int, float)):
                value = float(left) * float(right)
            else:
                return ""
        except Exception:
            return ""
    return max(str(value).splitlines() or [""])


def _auto_size_columns(sheet: Any, start_row: int = 12) -> None:
    merged: set[str] = set()
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col == max_col and min_row == max_row:
            continue
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                merged.add(f"{get_column_letter(column)}{row}")
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        maximum = 0
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if cell.coordinate in merged:
                continue
            maximum = max(maximum, len(_cell_text_for_width(cell)))
        if maximum:
            sheet.column_dimensions[letter].width = min(45, max(8, maximum + 2))
            sheet.column_dimensions[letter].bestFit = True


def export_yu_compact_workbook(
    *,
    template_path: str | Path,
    output_path: str | Path,
    order_date: str,
    order_number: str,
    item_numbers_with_qty: Iterable[tuple[str, Decimal | float | int]],
    worksheet_name: str = SHEET_NAME,
) -> tuple[tuple[str, int], ...]:
    path = Path(template_path)
    scan = scan_yu_workbook(path, worksheet_name=worksheet_name)
    grouped: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    original_number: dict[str, str] = {}
    for item_number, quantity in item_numbers_with_qty:
        key = clean_item_key(item_number)
        grouped[key] += Decimal(str(quantity))
        original_number.setdefault(key, str(item_number).strip())

    rows_with_qty: list[tuple[int, Decimal]] = []
    resolved_rows: list[tuple[str, int]] = []
    missing: list[str] = []
    duplicates: dict[str, list[int]] = {}
    for key, quantity in grouped.items():
        item_number = original_number[key]
        rows = _rows_for_item(scan, item_number)
        if not rows:
            missing.append(item_number)
        elif len(rows) > 1:
            duplicates[item_number] = rows
        else:
            rows_with_qty.append((rows[0], quantity))
            resolved_rows.append((item_number, rows[0]))
    if missing or duplicates:
        parts = []
        if missing:
            parts.append("Missing Column A mappings: " + ", ".join(sorted(missing)))
        if duplicates:
            parts.append(
                "Duplicate Column A mappings: "
                + "; ".join(
                    f"{item}: rows {', '.join(str(row) for row in rows)}"
                    for item, rows in sorted(duplicates.items())
                )
            )
        raise ValueError("\n".join(parts))

    try:
        from PIL import Image as _PILImage  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "The YU order logo cannot be exported because Pillow is not installed. "
            "Run the project dependency install and try again."
        ) from exc

    keep_vba = path.suffix.casefold() == ".xlsm"
    source_workbook = load_workbook(path, keep_vba=keep_vba)
    try:
        source = source_workbook[worksheet_name]
        source[DATE_CELL] = order_date
        source[ORDER_NUMBER_CELL] = order_number
        quantity_index = column_index_from_string(QUANTITY_COLUMN)
        quantity_by_row: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        for source_row, quantity in rows_with_qty:
            quantity_by_row[source_row] += quantity
        for row_number, quantity in quantity_by_row.items():
            source.cell(row_number, quantity_index).value = float(quantity)

        footer_start = int(scan["footer_start_row"])
        footer_end = int(scan["footer_end_row"])
        min_col = column_index_from_string(EXPORT_MIN_COLUMN)
        max_col = column_index_from_string(EXPORT_MAX_COLUMN)

        def row_has_item(row_number: int) -> bool:
            values = [source.cell(row_number, col).value for col in range(2, 9)]
            return _row_is_detail(values)

        def row_has_content(row_number: int) -> bool:
            return any(
                source.cell(row_number, col).value not in (None, "")
                for col in range(min_col, max_col + 1)
            )

        matched = set(quantity_by_row)
        rows_to_keep = set(range(1, HEADER_END_ROW + 1))
        rows_to_keep.update(range(footer_start, footer_end + 1))
        rows_to_keep.update(matched)
        for row_number in sorted(matched):
            rows_to_keep.update(
                _collect_header_note_rows(
                    row_number,
                    header_end_row=HEADER_END_ROW,
                    matched_rows=matched,
                    row_has_item=row_has_item,
                    row_has_content=row_has_content,
                )
            )
        kept_rows = sorted(rows_to_keep)
        plan = _row_plan(
            kept_rows,
            row_has_item,
            header_end_row=HEADER_END_ROW,
            footer_start_row=footer_start,
        )
        row_map: dict[int, int] = {}
        output_row = 1
        for old_row in plan:
            if old_row is None:
                output_row += 1
            else:
                row_map[old_row] = output_row
                output_row += 1

        output_workbook = Workbook()
        output = output_workbook.active
        output.title = worksheet_name
        output.sheet_view.showGridLines = source.sheet_view.showGridLines
        output.sheet_properties = copy(source.sheet_properties)
        output.page_margins = copy(source.page_margins)
        output.page_setup = copy(source.page_setup)
        output.print_options = copy(source.print_options)
        output.sheet_format = copy(source.sheet_format)

        for source_col in range(min_col, max_col + 1):
            output_col = source_col - min_col + 1
            source_letter = get_column_letter(source_col)
            output_letter = get_column_letter(output_col)
            source_dimension = source.column_dimensions[source_letter]
            output_dimension = output.column_dimensions[output_letter]
            output_dimension.width = source_dimension.width
            output_dimension.hidden = source_dimension.hidden
            output_dimension.bestFit = source_dimension.bestFit
            output_dimension.collapsed = source_dimension.collapsed
            output_dimension.outlineLevel = source_dimension.outlineLevel

        for old_row in plan:
            if old_row is None:
                continue
            new_row = row_map[old_row]
            source_row_dimension = source.row_dimensions[old_row]
            output_row_dimension = output.row_dimensions[new_row]
            output_row_dimension.height = source_row_dimension.height
            output_row_dimension.hidden = source_row_dimension.hidden
            output_row_dimension.outlineLevel = source_row_dimension.outlineLevel
            output_row_dimension.collapsed = source_row_dimension.collapsed

            for source_col in range(min_col, max_col + 1):
                output_col = source_col - min_col + 1
                source_cell = source.cell(old_row, source_col)
                output_cell = output.cell(new_row, output_col)
                instruction_row = (
                    HEADER_END_ROW < old_row < footer_start and not row_has_item(old_row)
                )
                if output_col == 1 and instruction_row:
                    output_cell.value = None
                elif isinstance(source_cell.value, str) and source_cell.value.startswith("="):
                    origin = f"{get_column_letter(source_col)}{old_row}"
                    destination = f"{get_column_letter(output_col)}{new_row}"
                    try:
                        output_cell.value = Translator(
                            source_cell.value, origin=origin
                        ).translate_formula(destination)
                    except Exception:
                        output_cell.value = source_cell.value
                else:
                    output_cell.value = source_cell.value
                if source_cell.has_style:
                    output_cell.font = copy(source_cell.font)
                    output_cell.fill = copy(source_cell.fill)
                    output_cell.border = copy(source_cell.border)
                    output_cell.alignment = copy(source_cell.alignment)
                    output_cell.number_format = source_cell.number_format
                    output_cell.protection = copy(source_cell.protection)
                if source_cell.hyperlink:
                    output_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    output_cell.comment = copy(source_cell.comment)

        for merged_range in source.merged_cells.ranges:
            range_min_col, min_row, range_max_col, max_row = merged_range.bounds
            if range_min_col < min_col or range_max_col > max_col:
                continue
            if not all(row in row_map for row in range(min_row, max_row + 1)):
                continue
            output.merge_cells(
                start_row=row_map[min_row],
                start_column=range_min_col - min_col + 1,
                end_row=row_map[max_row],
                end_column=range_max_col - min_col + 1,
            )

        header_images = []
        for image in getattr(source, "_images", []):
            anchor = getattr(image, "anchor", None)
            if anchor is None or not hasattr(anchor, "_from"):
                continue
            if anchor._from.row + 1 <= HEADER_END_ROW:
                header_images.append(image)

        if not header_images:
            raise RuntimeError(
                "The selected YU template did not expose a readable header logo. "
                "Confirm that the master workbook still contains the Windsor logo "
                "and that Pillow is installed."
            )

        copied_header_images = 0
        header_image_errors: list[str] = []
        for image in header_images:
            try:
                anchor = image.anchor
                old_row = anchor._from.row + 1
                old_col = anchor._from.col + 1
                target_row = row_map.get(old_row, old_row)
                image_data = image._data()
                new_image = XLImage(BytesIO(image_data))
                new_anchor = deepcopy(anchor)
                new_anchor._from.row = max(0, target_row - 1)
                new_anchor._from.col = max(
                    0,
                    min(old_col - min_col, max_col - min_col),
                )
                if hasattr(new_anchor, "to"):
                    old_to_row = anchor.to.row + 1
                    old_to_col = anchor.to.col + 1
                    target_to_row = row_map.get(old_to_row, old_to_row)
                    new_anchor.to.row = max(0, target_to_row - 1)
                    new_anchor.to.col = max(
                        0,
                        min(old_to_col - min_col, max_col - min_col),
                    )
                new_image.anchor = new_anchor
                output.add_image(new_image)
                copied_header_images += 1
            except Exception as exc:
                header_image_errors.append(f"{type(exc).__name__}: {exc}")

        if copied_header_images != len(header_images):
            detail = "; ".join(header_image_errors) or "unknown image-copy error"
            raise RuntimeError(
                "The YU workbook export was stopped because the header logo could "
                f"not be preserved. {detail}"
            )

        amount_source_column = column_index_from_string("M")
        if footer_end in row_map:
            amount_output_column = get_column_letter(amount_source_column - min_col + 1)
            total_output_row = row_map[footer_end]
            detail_output_rows = [
                row_map[row]
                for row in kept_rows
                if HEADER_END_ROW < row < footer_start and row_has_item(row)
            ]
            if detail_output_rows:
                output[f"{amount_output_column}{total_output_row}"] = (
                    f"=SUM({amount_output_column}{min(detail_output_rows)}:"
                    f"{amount_output_column}{max(detail_output_rows)})"
                )

        _auto_size_columns(output, start_row=12)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook.save(output_path)
    finally:
        source_workbook.close()
    return tuple(sorted(resolved_rows, key=lambda value: value[0].casefold()))


def _safe_filename_component(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    return cleaned.strip("._") or "YU_ORDER"


def export_yu_manufacture_order(
    session: Session,
    *,
    order_id: uuid.UUID,
    template_path: str | Path,
    output_directory: str | Path,
    expected_mtime_ns: int,
    worksheet_name: str = SHEET_NAME,
) -> YUExportResult:
    order = _load_order(session, order_id)
    path = Path(template_path)
    if path.stat().st_mtime_ns != int(expected_mtime_ns):
        raise YUWorkbookChanged(
            "The YU workbook changed after validation. Revalidate before exporting."
        )
    report = validate_yu_order(
        session,
        order_id=order_id,
        template_path=path,
        worksheet_name=worksheet_name,
    )
    if not report.ready:
        raise ValueError(
            "The YU order cannot be exported until every line has one unique Column A mapping."
        )
    order_lines = [
        (
            line.item.item_number,
            Decimal(line.ordered_quantity) - Decimal(line.cancelled_quantity),
        )
        for line in order.lines
        if Decimal(line.ordered_quantity) - Decimal(line.cancelled_quantity) > 0
    ]
    if not order_lines:
        raise ValueError("There are no positive open quantities to export.")
    output_dir = Path(output_directory) / "YU Orders"
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_order = _safe_filename_component(order.order_number)
    output_path = output_dir / f"yuchang_order_{safe_order}.xlsx"
    audit_path = output_dir / f"yuchang_order_{safe_order}_audit.csv"
    if output_path.exists() or audit_path.exists():
        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"yuchang_order_{safe_order}_{suffix}.xlsx"
        audit_path = output_dir / f"yuchang_order_{safe_order}_{suffix}_audit.csv"
    resolved = export_yu_compact_workbook(
        template_path=path,
        output_path=output_path,
        order_date=order.order_date.strftime("%d/%m/%Y"),
        order_number=order.order_number,
        item_numbers_with_qty=order_lines,
        worksheet_name=worksheet_name,
    )
    resolved_map = dict(resolved)
    with audit_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "order_number",
                "item_number",
                "quantity",
                "resolved_workbook_row",
                "template_path",
                "exported_at",
            ]
        )
        for item_number, quantity in order_lines:
            writer.writerow(
                [
                    order.order_number,
                    item_number,
                    str(quantity),
                    resolved_map.get(item_number, ""),
                    str(path),
                    utc_now().isoformat(timespec="seconds"),
                ]
            )
    return YUExportResult(
        output_path=str(output_path),
        audit_path=str(audit_path),
        item_count=len(resolved),
        line_count=len(order_lines),
        workbook_mtime_ns=int(path.stat().st_mtime_ns),
        resolved_rows=resolved,
    )


def add_yu_export_audit(
    session: Session,
    *,
    order_id: uuid.UUID,
    actor_user_id: uuid.UUID,
    result: YUExportResult,
) -> None:
    actor = _actor(session, actor_user_id)
    order = session.get(ManufactureOrder, order_id)
    if order is None:
        raise LookupError("Manufacture order not found.")
    session.add(
        AuditEvent(
            actor_user_id=actor.user_id,
            action="manufacture_order.yu_exported",
            entity_type="manufacture_order",
            entity_id=str(order.manufacture_order_id),
            source="web",
            summary=(
                f"Exported YU order {order.order_number}: {result.item_count} item(s), "
                f"{result.line_count} line(s)."
            )[:500],
            after_json=json.dumps(
                {
                    "output_path": result.output_path,
                    "audit_path": result.audit_path,
                    "resolved_rows": list(result.resolved_rows),
                },
                sort_keys=True,
            ),
        )
    )
    session.flush()
