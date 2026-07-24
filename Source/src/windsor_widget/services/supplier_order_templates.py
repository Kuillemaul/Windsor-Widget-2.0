"""Supplier order-form template discovery and audited configuration."""

from __future__ import annotations

import json
import os
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, true
from sqlalchemy.orm import Session

from windsor_widget.db.models import AppUser, AuditEvent, Supplier, SupplierOrderTemplate
from windsor_widget.db.models.audit import utc_now

YUCHANG_TEMPLATE_KIND = "yuchang_compact_xlsx"
DEFAULT_YUCHANG_TEMPLATE_FOLDER = (
    "C:\\Users\\WindsorTradingInfo\\WINDSOR TRADING CO TRUST\\"
    "Windsor Trading - Documents (1)\\data\\Excel\\Suppliers\\Order forms\\"
    "Yuchang - Orders"
)


@dataclass(frozen=True, slots=True)
class TemplateFileOption:
    file_name: str
    full_path: str
    modified_at: datetime
    size_bytes: int


@dataclass(frozen=True, slots=True)
class SupplierTemplateView:
    template_id: uuid.UUID | None
    supplier_id: uuid.UUID
    supplier_name: str
    folder_path: str
    file_name: str
    full_path: str | None
    worksheet_name: str
    configured: bool
    file_exists: bool
    verified_at: datetime | None
    verified_by: str | None
    available_files: tuple[TemplateFileOption, ...]
    message: str | None


def is_yuchang_supplier_name(value: str) -> bool:
    normalized = "".join(ch for ch in str(value or "").casefold() if ch.isalnum())
    return "yuchang" in normalized


def _active_user(session: Session, actor_user_id: uuid.UUID) -> AppUser:
    actor = session.get(AppUser, actor_user_id)
    if actor is None or not actor.is_active:
        raise LookupError("The signed-in user is no longer active.")
    return actor


def _clean_folder(value: str) -> Path:
    text = str(value or "").strip().strip('"')
    if not text:
        raise ValueError("Template folder is required.")
    return Path(text).expanduser()


def _safe_file_name(value: str) -> str:
    name = Path(str(value or "").strip()).name
    if not name:
        raise ValueError("Select a workbook file.")
    if name != str(value or "").strip():
        raise ValueError("Workbook must be selected from the configured folder.")
    if Path(name).suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise ValueError("The YU template must be an .xlsx or .xlsm workbook.")
    if name.startswith("~$"):
        raise ValueError("Excel temporary files cannot be selected as templates.")
    return name


def list_template_files(folder_path: str) -> tuple[TemplateFileOption, ...]:
    if not str(folder_path or "").strip():
        return ()
    folder = Path(str(folder_path).strip().strip('"')).expanduser()
    if not folder.exists() or not folder.is_dir():
        return ()
    options: list[TemplateFileOption] = []
    for path in folder.iterdir():
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            continue
        try:
            stat = path.stat()
        except OSError:
            continue
        options.append(
            TemplateFileOption(
                file_name=path.name,
                full_path=str(path),
                modified_at=datetime.fromtimestamp(stat.st_mtime),
                size_bytes=int(stat.st_size),
            )
        )
    return tuple(
        sorted(options, key=lambda value: (value.modified_at, value.file_name), reverse=True)
    )


def _validate_yuchang_workbook(path: Path, worksheet_name: str = "Sheet1") -> None:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"YU workbook was not found: {path}")
    keep_vba = path.suffix.casefold() == ".xlsm"
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=keep_vba,
        )
    except PermissionError as exc:
        raise ValueError(
            "Windows denied access to the selected YU workbook. "
            "Close the workbook in Excel, ensure the SharePoint/OneDrive file "
            "is fully downloaded using 'Always keep on this device', then try again. "
            f"Workbook: {path}"
        ) from exc

    try:
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {worksheet_name!r} was not found. Available sheets: "
                + ", ".join(workbook.sheetnames)
            )
        sheet = workbook[worksheet_name]
        heading_text = " ".join(
            str(sheet.cell(row, column).value or "")
            for row in range(1, 13)
            for column in range(1, min(int(sheet.max_column or 1), 14) + 1)
        )
        if "hengchang textile factory" not in heading_text.casefold():
            raise ValueError(
                "The selected workbook does not contain the expected "
                "Hengchang Textile Factory heading in the order-form header."
            )
        header_values = {
            str(sheet.cell(12, column).value or "").strip().casefold()
            for column in range(1, 14)
        }
        if "our number" not in header_values or "qty" not in header_values:
            raise ValueError(
                "The selected workbook does not look like the current YU order form."
            )
    finally:
        workbook.close()


def get_supplier_template_view(
    session: Session,
    supplier_id: uuid.UUID,
) -> SupplierTemplateView:
    supplier = session.get(Supplier, supplier_id)
    if supplier is None:
        raise LookupError("Supplier not found.")
    row = session.scalar(
        select(SupplierOrderTemplate).where(
            SupplierOrderTemplate.supplier_id == supplier_id,
            SupplierOrderTemplate.template_kind == YUCHANG_TEMPLATE_KIND,
            SupplierOrderTemplate.is_active == true(),
        )
    )
    folder = row.folder_path if row is not None else DEFAULT_YUCHANG_TEMPLATE_FOLDER
    file_name = row.file_name if row is not None else ""
    full_path = str(Path(folder) / file_name) if file_name else None
    exists = bool(full_path and Path(full_path).exists())
    message = None
    if row is None:
        message = "Select the active YU workbook before validating this order."
    elif not exists:
        message = "The configured workbook cannot currently be found on this PC."
    return SupplierTemplateView(
        template_id=row.supplier_order_template_id if row is not None else None,
        supplier_id=supplier.supplier_id,
        supplier_name=supplier.display_name,
        folder_path=folder,
        file_name=file_name,
        full_path=full_path,
        worksheet_name=row.worksheet_name if row is not None else "Sheet1",
        configured=row is not None,
        file_exists=exists,
        verified_at=row.verified_at if row is not None else None,
        verified_by=(
            row.verified_by.display_name
            if row is not None and row.verified_by is not None
            else None
        ),
        available_files=list_template_files(folder),
        message=message,
    )


def save_supplier_template(
    session: Session,
    *,
    supplier_id: uuid.UUID,
    folder_path: str,
    file_name: str,
    actor_user_id: uuid.UUID,
    worksheet_name: str = "Sheet1",
) -> SupplierOrderTemplate:
    actor = _active_user(session, actor_user_id)
    supplier = session.get(Supplier, supplier_id)
    if supplier is None or not supplier.is_active:
        raise LookupError("Active supplier not found.")
    if not is_yuchang_supplier_name(supplier.display_name):
        raise ValueError("The YU order-form template can only be assigned to Yuchang.")

    folder = _clean_folder(folder_path)
    name = _safe_file_name(file_name)
    target = folder / name
    try:
        resolved_folder = folder.resolve(strict=True)
        resolved_target = target.resolve(strict=True)
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Template folder or workbook was not found: {target}") from exc
    if resolved_target.parent != resolved_folder:
        raise ValueError("Workbook must be inside the configured template folder.")
    _validate_yuchang_workbook(resolved_target, worksheet_name=worksheet_name)

    row = session.scalar(
        select(SupplierOrderTemplate).where(
            SupplierOrderTemplate.supplier_id == supplier_id,
            SupplierOrderTemplate.template_kind == YUCHANG_TEMPLATE_KIND,
        )
    )
    before = None
    if row is None:
        row = SupplierOrderTemplate(
            supplier_id=supplier_id,
            template_kind=YUCHANG_TEMPLATE_KIND,
            folder_path=str(resolved_folder),
            file_name=name,
            worksheet_name=worksheet_name,
            is_active=True,
        )
        session.add(row)
        action = "supplier_order_template.created"
    else:
        before = {
            "folder_path": row.folder_path,
            "file_name": row.file_name,
            "worksheet_name": row.worksheet_name,
            "is_active": row.is_active,
        }
        row.folder_path = str(resolved_folder)
        row.file_name = name
        row.worksheet_name = worksheet_name
        row.is_active = True
        action = "supplier_order_template.updated"

    row.verified_at = utc_now()
    row.verified_by_user_id = actor.user_id
    row.updated_at = utc_now()
    session.flush()
    session.add(
        AuditEvent(
            actor_user_id=actor.user_id,
            action=action,
            entity_type="supplier_order_template",
            entity_id=str(row.supplier_order_template_id),
            source="web",
            summary=(
                f"Configured {supplier.display_name} order template: {resolved_target}."
            )[:500],
            before_json=json.dumps(before, sort_keys=True) if before else None,
            after_json=json.dumps(
                {
                    "supplier_id": str(supplier_id),
                    "folder_path": str(resolved_folder),
                    "file_name": name,
                    "worksheet_name": worksheet_name,
                },
                sort_keys=True,
            ),
        )
    )
    session.flush()
    return row


def open_template_on_server(path: str) -> None:
    target = Path(str(path or "").strip())
    if not target.exists():
        raise FileNotFoundError(f"YU workbook was not found: {target}")
    if os.name != "nt":
        raise RuntimeError("Opening the workbook is only supported on the Windows server PC.")
    os.startfile(str(target))  # type: ignore[attr-defined]
