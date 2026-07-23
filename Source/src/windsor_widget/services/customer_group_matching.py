# Review-first customer grouping and price-file matching.

from __future__ import annotations

import csv
import difflib
import json
import ntpath
import re
import uuid
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, true
from sqlalchemy.orm import Session

from windsor_widget.db.models import (
    AuditEvent,
    CustomerAccount,
    CustomerGroup,
    CustomerPriceFile,
)
from windsor_widget.db.models.audit import utc_now
from windsor_widget.imports.normalization import normalize_name

_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_IGNORED_FILE_KEYS = {"template", "1 template"}
_LEGAL_KEY_TOKENS = {"pty", "ltd", "limited", "pl", "p", "l", "co", "company", "inc"}
_STATE_PATTERNS = (
    r"\bn\s*\.?\s*s\s*\.?\s*w\s*\.?\b",
    r"\bv\s*\.?\s*i\s*\.?\s*c\s*\.?\b",
    r"\bq\s*\.?\s*l\s*\.?\s*d\s*\.?\b",
    r"\bs\s*\.?\s*a\s*\.?\b",
    r"\bw\s*\.?\s*a\s*\.?\b",
    r"\bt\s*\.?\s*a\s*\.?\s*s\s*\.?\b",
    r"\bn\s*\.?\s*t\s*\.?\b",
    r"\ba\s*\.?\s*c\s*\.?\s*t\s*\.?\b",
    r"\bnew\s+south\s+wales\b",
    r"\bvictoria\b",
    r"\bqueensland\b",
    r"\bsouth\s+australia\b",
    r"\bwestern\s+australia\b",
    r"\bwest\s+australia\b",
    r"\btasmania\b",
    r"\bnorthern\s+territory\b",
    r"\baustralian\s+capital\s+territory\b",
)


@dataclass(frozen=True, slots=True)
class WorkbookCustomer:
    name: str
    normalized_name: str
    group_key: str
    group_name: str
    city: str
    state: str


@dataclass(frozen=True, slots=True)
class PriceCandidate:
    path: str
    file_name: str
    key: str


@dataclass(frozen=True, slots=True)
class GroupProposal:
    group_key: str
    group_name: str
    account_ids: tuple[uuid.UUID, ...]
    account_names: tuple[str, ...]
    price_path: str | None
    price_file_name: str | None
    price_confidence: int | None
    price_method: str | None
    notes: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class GroupPlan:
    source_workbook: str
    proposals: tuple[GroupProposal, ...]
    matched_accounts: int
    unmatched_customer_names: tuple[str, ...]
    ambiguous_customer_names: tuple[str, ...]
    groups_without_price_file: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ApplySummary:
    groups_created: int
    groups_reused: int
    accounts_assigned: int
    accounts_already_correct: int
    accounts_skipped_existing_approved: int
    price_files_created: int
    price_files_reused: int


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _strip_archive_prefix(value: str) -> str:
    return re.sub(r"^\s*z{4,}\s*", "", value, flags=re.IGNORECASE)


def _strip_states(value: str) -> str:
    result = value
    for pattern in _STATE_PATTERNS:
        result = re.sub(pattern, " ", result, flags=re.IGNORECASE)
    result = re.sub(r"\(\s*\)", " ", result)
    result = re.sub(r"\s*[-–—,/]\s*$", "", result)
    return re.sub(r"\s+", " ", result).strip(" -–—,/")


def canonical_group_key(value: str) -> str:
    cleaned = _strip_states(_strip_archive_prefix(value))
    return " ".join(
        token
        for token in normalize_name(cleaned).split()
        if token not in _LEGAL_KEY_TOKENS
    )


def group_display_name(value: str) -> str:
    return _strip_states(_strip_archive_prefix(value)) or _strip_archive_prefix(value)


def _read_workbook(source: Path) -> tuple[tuple[WorkbookCustomer, ...], tuple[PriceCandidate, ...]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    if "Customer list Full" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'Customer list Full' sheet.")
    if "FILES" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'FILES' sheet.")

    customers: list[WorkbookCustomer] = []
    sheet = workbook["Customer list Full"]
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        name = _text(row[0] if row else None)
        if not name:
            continue
        customers.append(
            WorkbookCustomer(
                name=name,
                normalized_name=normalize_name(name),
                group_key=canonical_group_key(name),
                group_name=group_display_name(name),
                city=_text(row[2] if len(row) > 2 else None),
                state=_text(row[3] if len(row) > 3 else None),
            )
        )

    files: list[PriceCandidate] = []
    seen: set[str] = set()
    for row in workbook["FILES"].iter_rows(values_only=True):
        path = _text(row[0] if row else None)
        if not path:
            continue
        win_path = path.replace("/", "\\")
        if "\\old\\" in win_path.casefold():
            continue
        extension = ntpath.splitext(win_path)[1].casefold()
        if extension not in _ALLOWED_EXTENSIONS:
            continue
        file_name = ntpath.basename(win_path)
        stem = ntpath.splitext(file_name)[0].strip()
        if normalize_name(stem) in _IGNORED_FILE_KEYS:
            continue
        path_key = win_path.casefold()
        if path_key in seen:
            continue
        seen.add(path_key)
        files.append(
            PriceCandidate(
                path=path,
                file_name=file_name,
                key=canonical_group_key(stem),
            )
        )

    workbook.close()
    return tuple(customers), tuple(files)


def _resolve_account(
    candidates: tuple[CustomerAccount, ...],
    workbook_customer: WorkbookCustomer,
) -> CustomerAccount | None:
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        return None

    state = normalize_name(workbook_customer.state)
    city = normalize_name(workbook_customer.city)
    narrowed = tuple(
        account
        for account in candidates
        if (state and state == normalize_name(account.state))
        or (city and city == normalize_name(account.city))
    )
    return narrowed[0] if len(narrowed) == 1 else None


def _similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if len(left) >= 6 and (left in right or right in left):
        return 0.95
    direct = difflib.SequenceMatcher(None, left, right).ratio()
    token = difflib.SequenceMatcher(
        None,
        " ".join(sorted(left.split())),
        " ".join(sorted(right.split())),
    ).ratio()
    return max(direct, token)


def _select_price_file(
    group_key: str,
    account_names: tuple[str, ...],
    files: tuple[PriceCandidate, ...],
) -> tuple[PriceCandidate | None, int | None, str | None, str | None]:
    exact = [candidate for candidate in files if candidate.key == group_key]
    if exact:
        exact.sort(
            key=lambda candidate: (
                0 if ntpath.splitext(candidate.file_name)[1].casefold() in {".xlsx", ".xlsm"} else 1,
                len(candidate.file_name),
                candidate.file_name.casefold(),
            )
        )
        note = (
            f"{len(exact)} exact current files matched; preferred {exact[0].file_name}."
            if len(exact) > 1
            else None
        )
        return exact[0], 100, "exact_group_name", note

    account_keys = {canonical_group_key(name) for name in account_names}
    account_exact = [candidate for candidate in files if candidate.key in account_keys]
    if account_exact:
        account_exact.sort(key=lambda candidate: (len(candidate.file_name), candidate.file_name.casefold()))
        return account_exact[0], 98, "exact_account_name", None

    tokens = {token for token in group_key.split() if len(token) >= 4}
    pool = [candidate for candidate in files if tokens.intersection(candidate.key.split())]
    scored = sorted(
        ((_similarity(group_key, candidate.key), candidate) for candidate in pool),
        key=lambda pair: (-pair[0], pair[1].file_name.casefold()),
    )
    if not scored:
        return None, None, None, None

    best_score, best = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    if best_score >= 0.90 and best_score - second_score >= 0.06:
        return best, int(round(best_score * 100)), "fuzzy_unique", None
    return None, None, None, (
        f"Best candidate was {best.file_name} at {best_score:.0%}; left unlinked."
    )


def build_group_plan(
    session: Session,
    source_workbook: str | Path,
    *,
    include_inactive: bool = False,
) -> GroupPlan:
    source = Path(source_workbook)
    if not source.is_file():
        raise FileNotFoundError(source)

    workbook_customers, files = _read_workbook(source)
    account_statement = select(CustomerAccount)
    if not include_inactive:
        account_statement = account_statement.where(CustomerAccount.is_active == true())
    accounts = tuple(session.scalars(account_statement))

    by_name: dict[str, list[CustomerAccount]] = {}
    for account in accounts:
        by_name.setdefault(account.normalized_name, []).append(account)

    grouped: dict[str, list[tuple[CustomerAccount, WorkbookCustomer]]] = {}
    matched_ids: set[uuid.UUID] = set()
    unmatched: list[str] = []
    ambiguous: list[str] = []

    for source_customer in workbook_customers:
        candidates = tuple(by_name.get(source_customer.normalized_name, ()))
        account = _resolve_account(candidates, source_customer)
        if account is None:
            (ambiguous if candidates else unmatched).append(source_customer.name)
            continue
        if account.customer_account_id in matched_ids:
            continue
        matched_ids.add(account.customer_account_id)
        grouped.setdefault(source_customer.group_key, []).append((account, source_customer))

    proposals: list[GroupProposal] = []
    no_file: list[str] = []

    for group_key, pairs in grouped.items():
        pairs.sort(key=lambda pair: pair[0].display_name.casefold())
        account_ids = tuple(pair[0].customer_account_id for pair in pairs)
        account_names = tuple(pair[0].display_name for pair in pairs)
        group_name = (
            min(
                (pair[1].group_name for pair in pairs),
                key=lambda value: (len(value), value.casefold()),
            )
            if len(pairs) > 1
            else pairs[0][0].display_name
        )
        price_file, confidence, method, note = _select_price_file(
            group_key, account_names, files
        )
        notes = []
        if len(pairs) > 1:
            notes.append(f"Combined {len(pairs)} accounts.")
        if note:
            notes.append(note)
        if price_file is None:
            no_file.append(group_name)

        proposals.append(
            GroupProposal(
                group_key=group_key,
                group_name=group_name,
                account_ids=account_ids,
                account_names=account_names,
                price_path=price_file.path if price_file else None,
                price_file_name=price_file.file_name if price_file else None,
                price_confidence=confidence,
                price_method=method,
                notes=tuple(notes),
            )
        )

    proposals.sort(key=lambda proposal: proposal.group_name.casefold())
    return GroupPlan(
        source_workbook=str(source),
        proposals=tuple(proposals),
        matched_accounts=len(matched_ids),
        unmatched_customer_names=tuple(sorted(set(unmatched), key=str.casefold)),
        ambiguous_customer_names=tuple(sorted(set(ambiguous), key=str.casefold)),
        groups_without_price_file=tuple(sorted(set(no_file), key=str.casefold)),
    )


def apply_group_plan(
    session: Session,
    plan: GroupPlan,
    *,
    actor_user_id: uuid.UUID,
) -> ApplySummary:
    groups = {
        group.normalized_name: group
        for group in session.scalars(select(CustomerGroup))
    }
    accounts = {
        account.customer_account_id: account
        for account in session.scalars(select(CustomerAccount))
    }

    created = reused = assigned = already = skipped = files_created = files_reused = 0

    for proposal in plan.proposals:
        group = groups.get(proposal.group_key)
        if group is None:
            group = CustomerGroup(
                display_name=proposal.group_name,
                normalized_name=proposal.group_key,
                is_active=True,
                notes=f"Created from {Path(plan.source_workbook).name}.",
            )
            session.add(group)
            session.flush()
            groups[proposal.group_key] = group
            created += 1
            session.add(
                AuditEvent(
                    actor_user_id=actor_user_id,
                    action="customer.group.created",
                    entity_type="customer_group",
                    entity_id=str(group.customer_group_id),
                    source="customer_group_import",
                    summary=f"Created customer group {group.display_name}.",
                    after_json=json.dumps(
                        {
                            "display_name": group.display_name,
                            "normalized_name": group.normalized_name,
                        },
                        sort_keys=True,
                    ),
                )
            )
        else:
            reused += 1

        for account_id in proposal.account_ids:
            account = accounts[account_id]
            if account.customer_group_id == group.customer_group_id:
                account.group_match_status = "approved"
                already += 1
                continue
            if account.customer_group_id is not None and account.group_match_status == "approved":
                skipped += 1
                continue

            before = {
                "customer_group_id": str(account.customer_group_id) if account.customer_group_id else None,
                "group_match_status": account.group_match_status,
            }
            account.customer_group_id = group.customer_group_id
            account.group_match_status = "approved"
            assigned += 1
            session.add(
                AuditEvent(
                    actor_user_id=actor_user_id,
                    action="customer.group.assigned",
                    entity_type="customer_account",
                    entity_id=str(account.customer_account_id),
                    source="customer_group_import",
                    summary=f"{account.display_name} assigned to {group.display_name}.",
                    before_json=json.dumps(before, sort_keys=True),
                    after_json=json.dumps(
                        {
                            "customer_group_id": str(group.customer_group_id),
                            "group_name": group.display_name,
                            "group_match_status": "approved",
                        },
                        sort_keys=True,
                    ),
                )
            )

        if proposal.price_path and proposal.price_file_name:
            existing = session.scalar(
                select(CustomerPriceFile).where(
                    CustomerPriceFile.customer_group_id == group.customer_group_id,
                    CustomerPriceFile.file_path == proposal.price_path,
                )
            )
            if existing is None:
                existing = CustomerPriceFile(
                    customer_group_id=group.customer_group_id,
                    file_path=proposal.price_path,
                    file_name=proposal.price_file_name,
                    match_status="approved",
                    confidence=proposal.price_confidence,
                    is_active=True,
                    verified_at=utc_now(),
                    verified_by_user_id=actor_user_id,
                )
                session.add(existing)
                session.flush()
                files_created += 1
                session.add(
                    AuditEvent(
                        actor_user_id=actor_user_id,
                        action="customer.price_file.linked",
                        entity_type="customer_price_file",
                        entity_id=str(existing.customer_price_file_id),
                        source="customer_group_import",
                        summary=f"Linked {proposal.price_file_name} to {group.display_name}.",
                        after_json=json.dumps(
                            {
                                "customer_group_id": str(group.customer_group_id),
                                "file_path": proposal.price_path,
                                "confidence": proposal.price_confidence,
                                "match_method": proposal.price_method,
                            },
                            sort_keys=True,
                        ),
                    )
                )
            else:
                existing.is_active = True
                existing.match_status = "approved"
                existing.confidence = proposal.price_confidence
                existing.verified_at = utc_now()
                existing.verified_by_user_id = actor_user_id
                files_reused += 1

    session.flush()
    return ApplySummary(created, reused, assigned, already, skipped, files_created, files_reused)


def write_group_report(plan: GroupPlan, destination: str | Path) -> Path:
    output = Path(destination)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "group_name",
                "group_key",
                "account_count",
                "accounts",
                "price_file",
                "price_confidence",
                "price_match_method",
                "notes",
            ]
        )
        for proposal in plan.proposals:
            writer.writerow(
                [
                    proposal.group_name,
                    proposal.group_key,
                    len(proposal.account_ids),
                    " | ".join(proposal.account_names),
                    proposal.price_path or "",
                    proposal.price_confidence or "",
                    proposal.price_method or "",
                    " | ".join(proposal.notes),
                ]
            )
        writer.writerow([])
        writer.writerow(["UNMATCHED CUSTOMER NAMES"])
        for value in plan.unmatched_customer_names:
            writer.writerow([value])
        writer.writerow([])
        writer.writerow(["AMBIGUOUS CUSTOMER NAMES"])
        for value in plan.ambiguous_customer_names:
            writer.writerow([value])
        writer.writerow([])
        writer.writerow(["GROUPS WITHOUT SAFE PRICE FILE"])
        for value in plan.groups_without_price_file:
            writer.writerow([value])
    return output
