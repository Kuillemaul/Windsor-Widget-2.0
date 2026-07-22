"""Preview and commit immutable MYOB inventory snapshots.

The source is the MYOB ``Analyse Inventory [Summary]`` workbook containing:
Item No., Item Name, On Hand, Committed, On Order and Available.
"""

from __future__ import annotations

import hashlib
import json
import re
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import insert, select, true, update
from sqlalchemy.orm import Session

from windsor_widget.db.models import (
    AppUser,
    AuditEvent,
    InventorySnapshot,
    InventorySnapshotLine,
    Item,
)
from windsor_widget.db.models.audit import utc_now

_ZERO = Decimal("0")
_BALANCE_TOLERANCE = Decimal("0.000001")
_REQUIRED_FIELDS = (
    "item_number",
    "item_name",
    "on_hand",
    "committed",
    "on_order",
    "available",
)
_HEADER_ALIASES = {
    "item_number": {"itemno", "itemnumber"},
    "item_name": {"itemname", "description"},
    "on_hand": {"onhand", "qtyonhand", "quantityonhand"},
    "committed": {"committed", "qtycommitted", "quantitycommitted"},
    "on_order": {"onorder", "qtyonorder", "quantityonorder"},
    "available": {"available", "qtyavailable", "quantityavailable"},
}


class InventorySnapshotError(ValueError):
    """Raised when an inventory snapshot cannot be trusted or committed."""


@dataclass(frozen=True, slots=True)
class InventorySourceRow:
    source_row_number: int
    item_number: str
    item_name: str
    on_hand: Decimal
    committed: Decimal
    on_order: Decimal
    available: Decimal


@dataclass(frozen=True, slots=True)
class InventorySnapshotPreview:
    source_path: Path
    source_file_name: str
    source_sha256: str
    captured_at: datetime
    row_count: int
    matched_item_count: int
    unmatched_item_numbers: tuple[str, ...]
    total_on_hand: Decimal
    total_committed: Decimal
    total_on_order: Decimal
    total_available: Decimal
    already_imported: bool
    existing_snapshot_id: uuid.UUID | None


@dataclass(frozen=True, slots=True)
class InventorySnapshotCommit:
    mode: str
    inventory_snapshot_id: uuid.UUID
    captured_at: datetime
    source_file_name: str
    source_sha256: str
    row_count: int


def _normalise_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").casefold())


def _decimal(value: object, *, row_number: int, field_name: str) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        return _ZERO
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "").replace("$", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise InventorySnapshotError(
            f"Inventory row {row_number} has invalid {field_name} value {value!r}."
        ) from exc


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _naive_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(UTC).replace(tzinfo=None)


def _captured_at(workbook, path: Path, override: datetime | None) -> datetime:
    if override is not None:
        return _naive_utc(override)
    modified = getattr(workbook.properties, "modified", None)
    if isinstance(modified, datetime):
        return _naive_utc(modified)
    return datetime.fromtimestamp(path.stat().st_mtime, tz=UTC).replace(tzinfo=None)


def _header_map(values: tuple[object, ...]) -> dict[str, int] | None:
    normalised = [_normalise_header(value) for value in values]
    mapping: dict[str, int] = {}
    for field_name, aliases in _HEADER_ALIASES.items():
        for index, header in enumerate(normalised):
            if header in aliases:
                mapping[field_name] = index
                break
    return mapping if all(field in mapping for field in _REQUIRED_FIELDS) else None


def parse_inventory_workbook(
    source_path: str | Path,
    *,
    captured_at: datetime | None = None,
) -> tuple[tuple[InventorySourceRow, ...], datetime, str]:
    """Parse and internally reconcile an MYOB Analyse Inventory workbook."""

    path = Path(source_path).expanduser().resolve()
    if not path.is_file():
        raise InventorySnapshotError(f"Inventory workbook does not exist: {path}")
    if path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise InventorySnapshotError("Inventory snapshot source must be an .xlsx or .xlsm file.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        located = None
        for worksheet in workbook.worksheets:
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=100, values_only=True), start=1
            ):
                mapping = _header_map(tuple(row))
                if mapping is not None:
                    located = (worksheet, row_number, mapping)
                    break
            if located is not None:
                break
        if located is None:
            raise InventorySnapshotError(
                "Could not find the Analyse Inventory header row. Expected Item No., "
                "Item Name, On Hand, Committed, On Order and Available."
            )

        worksheet, header_row_number, mapping = located
        rows: list[InventorySourceRow] = []
        seen: dict[str, int] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            values = tuple(row)
            item_number = str(values[mapping["item_number"]] or "").strip()
            item_name = str(values[mapping["item_name"]] or "").strip()
            numeric = {
                field_name: _decimal(
                    values[column_index] if column_index < len(values) else None,
                    row_number=row_number,
                    field_name=field_name,
                )
                for field_name, column_index in mapping.items()
                if field_name not in {"item_number", "item_name"}
            }
            if not item_number and not item_name and all(value == 0 for value in numeric.values()):
                continue
            if not item_number:
                raise InventorySnapshotError(
                    f"Inventory row {row_number} contains quantities but no item number."
                )

            identity = item_number.casefold()
            if identity in seen:
                raise InventorySnapshotError(
                    f"Inventory item {item_number!r} appears on rows {seen[identity]} and "
                    f"{row_number}."
                )
            seen[identity] = row_number

            expected_available = (
                numeric["on_hand"] - numeric["committed"] + numeric["on_order"]
            )
            if abs(expected_available - numeric["available"]) > _BALANCE_TOLERANCE:
                raise InventorySnapshotError(
                    f"Inventory row {row_number} ({item_number}) does not balance: "
                    f"On Hand - Committed + On Order = {expected_available}, but Available "
                    f"is {numeric['available']}."
                )

            rows.append(
                InventorySourceRow(
                    source_row_number=row_number,
                    item_number=item_number,
                    item_name=item_name,
                    on_hand=numeric["on_hand"],
                    committed=numeric["committed"],
                    on_order=numeric["on_order"],
                    available=numeric["available"],
                )
            )

        if not rows:
            raise InventorySnapshotError("The inventory workbook contains no item rows.")
        return tuple(rows), _captured_at(workbook, path, captured_at), _sha256(path)
    finally:
        workbook.close()


def _item_ids(session: Session) -> dict[str, tuple[str, uuid.UUID]]:
    mapping: dict[str, tuple[str, uuid.UUID]] = {}
    for item_number, item_id in session.execute(select(Item.item_number, Item.item_id)):
        identity = item_number.casefold()
        if identity in mapping and mapping[identity][0] != item_number:
            raise InventorySnapshotError(
                "Item master contains case-insensitive duplicate item numbers: "
                f"{mapping[identity][0]!r} and {item_number!r}."
            )
        mapping[identity] = (item_number, item_id)
    return mapping


def preview_inventory_snapshot(
    session: Session,
    source_path: str | Path,
    *,
    captured_at: datetime | None = None,
) -> InventorySnapshotPreview:
    """Validate source rows and exact item linkage without changing the database."""

    path = Path(source_path).expanduser().resolve()
    rows, captured, source_hash = parse_inventory_workbook(path, captured_at=captured_at)
    item_ids = _item_ids(session)
    unmatched = tuple(
        row.item_number for row in rows if row.item_number.casefold() not in item_ids
    )
    existing = session.scalar(
        select(InventorySnapshot).where(InventorySnapshot.source_sha256 == source_hash)
    )
    return InventorySnapshotPreview(
        source_path=path,
        source_file_name=path.name,
        source_sha256=source_hash,
        captured_at=captured,
        row_count=len(rows),
        matched_item_count=len(rows) - len(unmatched),
        unmatched_item_numbers=unmatched,
        total_on_hand=sum((row.on_hand for row in rows), _ZERO),
        total_committed=sum((row.committed for row in rows), _ZERO),
        total_on_order=sum((row.on_order for row in rows), _ZERO),
        total_available=sum((row.available for row in rows), _ZERO),
        already_imported=existing is not None,
        existing_snapshot_id=(existing.inventory_snapshot_id if existing else None),
    )


def _bulk_insert_lines(
    session: Session,
    values: list[dict[str, object]],
    *,
    chunk_size: int = 1_000,
) -> None:
    for offset in range(0, len(values), chunk_size):
        session.execute(insert(InventorySnapshotLine), values[offset : offset + chunk_size])


def commit_inventory_snapshot(
    session: Session,
    source_path: str | Path,
    *,
    actor: AppUser,
    captured_at: datetime | None = None,
) -> InventorySnapshotCommit:
    """Commit one clean immutable snapshot and make it the sole current snapshot."""

    preview = preview_inventory_snapshot(session, source_path, captured_at=captured_at)
    if preview.unmatched_item_numbers:
        examples = ", ".join(preview.unmatched_item_numbers[:10])
        more = "" if len(preview.unmatched_item_numbers) <= 10 else " …"
        raise InventorySnapshotError(
            f"Inventory snapshot has {len(preview.unmatched_item_numbers)} unmatched item "
            f"number(s): {examples}{more}."
        )
    if preview.already_imported:
        existing = session.get(InventorySnapshot, preview.existing_snapshot_id)
        if existing is None:
            raise InventorySnapshotError("Existing inventory snapshot disappeared during import.")
        return InventorySnapshotCommit(
            mode="unchanged",
            inventory_snapshot_id=existing.inventory_snapshot_id,
            captured_at=existing.captured_at,
            source_file_name=existing.source_file_name,
            source_sha256=existing.source_sha256,
            row_count=existing.row_count,
        )

    rows, captured, source_hash = parse_inventory_workbook(
        preview.source_path, captured_at=preview.captured_at
    )
    item_ids = _item_ids(session)

    session.execute(
        update(InventorySnapshot)
        .where(InventorySnapshot.is_current == true())
        .values(is_current=False)
    )
    snapshot = InventorySnapshot(
        captured_at=captured,
        source_file_name=preview.source_file_name,
        source_sha256=source_hash,
        row_count=len(rows),
        is_current=True,
        committed_by_user_id=actor.user_id,
    )
    session.add(snapshot)
    session.flush()

    line_values: list[dict[str, object]] = []
    for row in rows:
        _, item_id = item_ids[row.item_number.casefold()]
        line_values.append(
            {
                "inventory_snapshot_line_id": uuid.uuid4(),
                "inventory_snapshot_id": snapshot.inventory_snapshot_id,
                "item_id": item_id,
                "source_row_number": row.source_row_number,
                "item_number_snapshot": row.item_number,
                "item_name_snapshot": row.item_name,
                "on_hand": row.on_hand,
                "committed": row.committed,
                "on_order": row.on_order,
                "available": row.available,
            }
        )
    _bulk_insert_lines(session, line_values)

    session.add(
        AuditEvent(
            actor_user_id=actor.user_id,
            action="inventory_snapshot_committed",
            entity_type="inventory_snapshot",
            entity_id=str(snapshot.inventory_snapshot_id),
            correlation_id=uuid.uuid4(),
            source="myob_inventory",
            summary=(
                f"Committed inventory snapshot {preview.source_file_name} with "
                f"{len(rows)} matched item rows."
            ),
            after_json=json.dumps(
                {
                    "captured_at": captured.isoformat(),
                    "source_file_name": preview.source_file_name,
                    "source_sha256": source_hash,
                    "row_count": len(rows),
                    "total_on_hand": str(preview.total_on_hand),
                    "total_committed": str(preview.total_committed),
                    "total_on_order": str(preview.total_on_order),
                    "total_available": str(preview.total_available),
                },
                ensure_ascii=False,
                sort_keys=True,
            ),
        )
    )
    session.flush()
    return InventorySnapshotCommit(
        mode="committed",
        inventory_snapshot_id=snapshot.inventory_snapshot_id,
        captured_at=captured,
        source_file_name=preview.source_file_name,
        source_sha256=source_hash,
        row_count=len(rows),
    )


def current_inventory_snapshot(session: Session) -> InventorySnapshot | None:
    return session.scalar(
        select(InventorySnapshot)
        .where(InventorySnapshot.is_current == true())
        .order_by(InventorySnapshot.captured_at.desc())
        .limit(1)
    )


def parse_iso_datetime(value: str) -> datetime:
    """Argparse-compatible ISO date/time parser."""

    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(
            f"Expected ISO date/time such as 2026-07-20T09:56:25, received {value!r}."
        ) from exc
    return _naive_utc(parsed)


def inventory_snapshot_age_days(snapshot: InventorySnapshot, *, now: datetime | None = None) -> int:
    reference = now or utc_now()
    return max(0, (reference.date() - snapshot.captured_at.date()).days)
