from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from windsor_widget.services.yu_order_export import (
    _rows_for_item,
    _write_mapping_cells,
    export_yu_compact_workbook,
    scan_yu_workbook,
)


def _workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B8"] = "To: - Hengchang Textile Factory"
    ws["B10"] = "Date: - "
    ws["G10"] = "Order No: - "
    headers = [
        "Our Number",
        "Item",
        "Size",
        "Colour",
        "Roll / ",
        "Mt /",
        "Labelled as",
        "Mt per",
        "Mt per",
        "No Of",
        "FOB Price",
        "Qty",
        "US $ ",
    ]
    for column, value in enumerate(headers, start=1):
        ws.cell(12, column).value = value
    ws["A15"] = "ITEM 001"
    ws["B15"] = "Test cord"
    ws["C15"] = "3 mm"
    ws["D15"] = "Black"
    ws["E15"] = "Spool"
    ws["F15"] = 100
    ws["G15"] = "TEST 3 MM"
    ws["K15"] = 0.25
    ws["L15"] = 0
    ws["M15"] = "=K15*L15"
    ws["A16"] = "ITEM002"
    ws["B16"] = "Test cord"
    ws["C16"] = "4 mm"
    ws["D16"] = "White"
    ws["E16"] = "Spool"
    ws["F16"] = 100
    ws["G16"] = "TEST 4 MM"
    ws["K16"] = 0.30
    ws["L16"] = 0
    ws["M16"] = "=K16*L16"
    ws["B20"] = "Method of Transport - BY SEA"
    ws["G24"] = "Order Total"
    ws["L24"] = "US $ "
    ws["M24"] = "=SUM(M15:M16)"
    review = wb.create_sheet("Match_Review")
    review["A4"] = "Source Row"
    review["B4"] = "Final Selection"
    review["C4"] = "Suggested Match"
    review["A5"] = 16
    review["C5"] = "ITEM002"
    logo_path = path.with_name(path.stem + "_logo.png")
    PILImage.new("RGB", (80, 24), "white").save(logo_path)
    logo = XLImage(logo_path)
    logo.anchor = "B1"
    ws.add_image(logo)
    wb.save(path)


def test_scan_uses_current_column_a_and_space_insensitive_key(tmp_path: Path) -> None:
    path = tmp_path / "yu.xlsx"
    _workbook(path)
    scan = scan_yu_workbook(path)
    assert _rows_for_item(scan, "ITEM001") == [15]
    assert _rows_for_item(scan, "ITEM002") == [16]
    assert scan["footer_start_row"] == 19
    assert scan["footer_end_row"] == 24


def test_mapping_update_changes_only_current_sheet1_column_a(tmp_path: Path) -> None:
    path = tmp_path / "yu.xlsx"
    _workbook(path)
    _write_mapping_cells(path, {16: "NEW002"}, worksheet_name="Sheet1")
    scan = scan_yu_workbook(path)
    assert _rows_for_item(scan, "NEW002") == [16]
    wb = load_workbook(path, data_only=False)
    try:
        assert wb["Sheet1"]["B16"].value == "Test cord"
        assert wb["Match_Review"]["C5"].value == "ITEM002"
    finally:
        wb.close()


def test_compact_export_preserves_heading_and_selected_quantity(tmp_path: Path) -> None:
    template = tmp_path / "yu.xlsx"
    output = tmp_path / "order.xlsx"
    _workbook(template)
    resolved = export_yu_compact_workbook(
        template_path=template,
        output_path=output,
        order_date="24/07/2026",
        order_number="YU-001",
        item_numbers_with_qty=[("ITEM002", 250)],
    )
    assert resolved == (("ITEM002", 16),)
    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["Sheet1"]
        assert wb.sheetnames == ["Sheet1"]
        assert len(getattr(ws, "_images", [])) == 1
        assert ws["B8"].value == "To: - Hengchang Textile Factory"
        assert ws["C10"].value == "24/07/2026"
        assert ws["H10"].value == "YU-001"
        exported_items = [
            ws.cell(row, 1).value
            for row in range(15, ws.max_row + 1)
            if ws.cell(row, 1).value
        ]
        assert exported_items == ["ITEM002"]
        assert ws["L15"].value == 250
        assert ws["M15"].value == "=K15*L15"
    finally:
        wb.close()
