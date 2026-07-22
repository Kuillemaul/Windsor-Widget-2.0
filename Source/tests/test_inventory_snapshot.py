from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from windsor_widget.db.base import Base
from windsor_widget.db.models import (
    AppUser,
    InventorySnapshot,
    InventorySnapshotLine,
    Item,
)
from windsor_widget.imports.inventory_snapshot import (
    InventorySnapshotError,
    commit_inventory_snapshot,
    preview_inventory_snapshot,
)


def _engine():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    return engine


def _workbook(
    path: Path,
    rows: list[tuple[object, ...]],
    *,
    modified: datetime = datetime(2026, 7, 19, 23, 56, 25),
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["B2"] = "Windsor Trading Australia Pty Ltd"
    worksheet["B7"] = "Analyse Inventory [Summary]"
    headers = ["Item No.", "Item Name", "On Hand", "Committed", "On Order", "Available"]
    for column, header in enumerate(headers, start=2):
        worksheet.cell(row=10, column=column, value=header)
    for row_number, row in enumerate(rows, start=12):
        for column, value in enumerate(row, start=2):
            worksheet.cell(row=row_number, column=column, value=value)
    workbook.properties.modified = modified
    workbook.save(path)
    workbook.close()


def _seed_items(session: Session) -> AppUser:
    actor = AppUser(username="brad", display_name="Brad Mayze")
    session.add_all(
        [
            actor,
            Item(
                item_number="I1",
                item_name="Item One",
                normalized_name="item one",
                is_active=True,
                is_inventoried=True,
            ),
            Item(
                item_number="I2",
                item_name="Item Two",
                normalized_name="item two",
                is_active=True,
                is_inventoried=True,
            ),
        ]
    )
    session.commit()
    return actor


def test_preview_and_commit_exact_balanced_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "zinvs1.xlsx"
    _workbook(
        source,
        [
            ("I1", "Item One", 100, 25, 10, 85),
            ("I2", "Item Two", 20, 30, 5, -5),
        ],
    )

    with Session(_engine()) as session:
        actor = _seed_items(session)
        captured = datetime(2026, 7, 19, 23, 56, 25)
        preview = preview_inventory_snapshot(session, source, captured_at=captured)
        assert preview.row_count == 2
        assert preview.matched_item_count == 2
        assert preview.unmatched_item_numbers == ()
        assert preview.total_on_hand == Decimal("120")
        assert preview.total_committed == Decimal("55")
        assert preview.total_on_order == Decimal("15")
        assert preview.total_available == Decimal("80")
        assert preview.captured_at == datetime(2026, 7, 19, 23, 56, 25)

        result = commit_inventory_snapshot(
            session, source, actor=actor, captured_at=captured
        )
        session.commit()
        assert result.mode == "committed"
        assert result.row_count == 2
        assert session.scalar(select(func.count(InventorySnapshot.inventory_snapshot_id))) == 1
        assert (
            session.scalar(
                select(func.count(InventorySnapshotLine.inventory_snapshot_line_id))
            )
            == 2
        )

        second = commit_inventory_snapshot(
            session, source, actor=actor, captured_at=captured
        )
        session.commit()
        assert second.mode == "unchanged"
        assert second.inventory_snapshot_id == result.inventory_snapshot_id
        assert session.scalar(select(func.count(InventorySnapshot.inventory_snapshot_id))) == 1


def test_new_snapshot_replaces_current_without_deleting_history(tmp_path: Path) -> None:
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"
    _workbook(first_path, [("I1", "Item One", 100, 20, 0, 80)])
    _workbook(
        second_path,
        [("I1", "Item One", 90, 10, 20, 100)],
        modified=datetime(2026, 7, 22, 1, 0, 0),
    )

    with Session(_engine()) as session:
        actor = _seed_items(session)
        first = commit_inventory_snapshot(session, first_path, actor=actor)
        session.commit()
        second = commit_inventory_snapshot(session, second_path, actor=actor)
        session.commit()

        snapshots = list(
            session.scalars(select(InventorySnapshot).order_by(InventorySnapshot.captured_at))
        )
        assert [snapshot.is_current for snapshot in snapshots] == [False, True]
        assert snapshots[0].inventory_snapshot_id == first.inventory_snapshot_id
        assert snapshots[1].inventory_snapshot_id == second.inventory_snapshot_id


def test_unmatched_item_blocks_commit(tmp_path: Path) -> None:
    source = tmp_path / "unmatched.xlsx"
    _workbook(source, [("MISSING", "Missing Item", 1, 0, 0, 1)])

    with Session(_engine()) as session:
        actor = _seed_items(session)
        preview = preview_inventory_snapshot(session, source)
        assert preview.unmatched_item_numbers == ("MISSING",)
        with pytest.raises(InventorySnapshotError, match="unmatched"):
            commit_inventory_snapshot(session, source, actor=actor)


def test_duplicate_and_unbalanced_rows_are_rejected(tmp_path: Path) -> None:
    duplicate = tmp_path / "duplicate.xlsx"
    _workbook(
        duplicate,
        [("I1", "Item One", 1, 0, 0, 1), ("i1", "Item One", 2, 0, 0, 2)],
    )
    unbalanced = tmp_path / "unbalanced.xlsx"
    _workbook(unbalanced, [("I1", "Item One", 10, 2, 1, 99)])

    with Session(_engine()) as session:
        _seed_items(session)
        with pytest.raises(InventorySnapshotError, match="appears on rows"):
            preview_inventory_snapshot(session, duplicate)
        with pytest.raises(InventorySnapshotError, match="does not balance"):
            preview_inventory_snapshot(session, unbalanced)
