"""Review-only match proposals built from approved source exports.

Nothing in this module writes to the database or turns a proposal into an approved
relationship. Exact, unique matches can be presented as high confidence; every
other match remains explicitly reviewable.
"""

from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path, PureWindowsPath

from openpyxl import load_workbook

from windsor_widget.imports.master_data import (
    CustomerMasterCandidate,
    ItemMasterCandidate,
    SupplierMasterCandidate,
)
from windsor_widget.imports.normalization import (
    clean_text,
    is_control_item_number,
    normalize_name,
    parse_date,
    parse_decimal,
)

SourceRow = Mapping[str, str | None]

_LEGAL_TOKENS = {
    "co",
    "company",
    "corp",
    "corporation",
    "inc",
    "incorporated",
    "limited",
    "ltd",
    "pty",
}
_NAME_NOISE_TOKENS = {
    "customer",
    "customers",
    "file",
    "files",
    "list",
    "price",
    "prices",
    "pricing",
    "the",
}
_JOINER_TOKENS = {"and", "of"}
_STATE_SUFFIXES = (
    ("south", "australia"),
    ("western", "australia"),
    ("west", "australia"),
    ("new", "zealand"),
    ("queensland",),
    ("tasmania",),
    ("victoria",),
    ("australian", "capital", "territory"),
    ("new", "south", "wales"),
    ("northern", "territory"),
    ("a", "c", "t"),
    ("n", "s", "w"),
    ("act",),
    ("nsw",),
    ("nt",),
    ("nz",),
    ("qld",),
    ("sa",),
    ("tas",),
    ("vic",),
    ("wa",),
)
_DISPLAY_STATE_SUFFIX = re.compile(
    r"\s*-\s*(?:A\.?C\.?T\.?|N\.?S\.?W\.?|N\.?T\.?|N\.?Z\.?|Q\.?L\.?D\.?|"
    r"S\.?A\.?|T\.?A\.?S\.?|V\.?I\.?C\.?|W\.?A\.?|Australian Capital Territory|"
    r"New South Wales|Northern Territory|New Zealand|Queensland|South Australia|"
    r"Tasmania|Victoria|West(?:ern)? Australia)\s*$",
    flags=re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class CustomerGroupProposal:
    group_key: str
    display_name: str
    account_record_ids: tuple[str, ...]
    account_names: tuple[str, ...]
    score: int
    method: str
    evidence: tuple[str, ...]
    requires_review: bool


@dataclass(frozen=True, slots=True)
class PriceFileReference:
    path: str
    file_name: str
    normalized_stem: str


@dataclass(frozen=True, slots=True)
class MatchAlternative:
    target_key: str
    target_name: str
    score: int


@dataclass(frozen=True, slots=True)
class CustomerPriceFileProposal:
    file_path: str
    file_name: str
    customer_group_key: str | None
    customer_group_name: str | None
    score: int
    method: str
    evidence: tuple[str, ...]
    alternatives: tuple[MatchAlternative, ...]
    requires_review: bool


@dataclass(frozen=True, slots=True)
class PurchaseEvidence:
    item_number: str
    supplier_name: str
    purchase_number: str | None
    purchase_date: date
    purchase_status: str | None
    price_text: str | None
    currency_code: str | None


@dataclass(frozen=True, slots=True)
class ItemSupplierProposal:
    item_number: str
    source_supplier_name: str
    supplier_record_id: str | None
    supplier_name: str | None
    score: int
    method: str
    evidence: tuple[str, ...]
    alternatives: tuple[MatchAlternative, ...]
    last_purchase: PurchaseEvidence | None
    requires_review: bool


def _drop_suffix(tokens: list[str], suffix: tuple[str, ...]) -> bool:
    if len(tokens) < len(suffix) or tuple(tokens[-len(suffix) :]) != suffix:
        return False
    del tokens[-len(suffix) :]
    return True


def _significant_tokens(value: str | None, *, strip_state_suffix: bool = False) -> tuple[str, ...]:
    tokens = normalize_name(value).split()
    if strip_state_suffix:
        for suffix in _STATE_SUFFIXES:
            if _drop_suffix(tokens, suffix):
                break
    return tuple(
        token
        for token in tokens
        if token not in _LEGAL_TOKENS
        and token not in _NAME_NOISE_TOKENS
        and token not in _JOINER_TOKENS
    )


def _signature(value: str | None, *, strip_state_suffix: bool = False) -> str:
    return " ".join(
        sorted(set(_significant_tokens(value, strip_state_suffix=strip_state_suffix)))
    )


def _display_group_name(account_names: Iterable[str]) -> str:
    stripped = [
        _DISPLAY_STATE_SUFFIX.sub("", name).strip() for name in account_names if name.strip()
    ]
    return min(stripped, key=lambda value: (len(value), value.casefold()))


def propose_customer_groups(
    customers: Iterable[CustomerMasterCandidate],
) -> tuple[CustomerGroupProposal, ...]:
    """Group accounts only when their normalized business-family names are exact."""

    grouped: dict[str, list[CustomerMasterCandidate]] = defaultdict(list)
    for customer in customers:
        group_key = _signature(customer.display_name, strip_state_suffix=True)
        if not group_key:
            group_key = customer.normalized_name
        grouped[group_key].append(customer)

    proposals: list[CustomerGroupProposal] = []
    for group_key, accounts in grouped.items():
        ordered = sorted(
            accounts,
            key=lambda account: (
                account.display_name.casefold(),
                account.myob_record_id or "",
            ),
        )
        account_names = tuple(account.display_name for account in ordered)
        record_ids = tuple(
            account.myob_record_id or f"name:{account.normalized_name}" for account in ordered
        )
        proposals.append(
            CustomerGroupProposal(
                group_key=group_key,
                display_name=_display_group_name(account_names),
                account_record_ids=record_ids,
                account_names=account_names,
                score=100,
                method="exact_normalized_business_family",
                evidence=(
                    f"All {len(ordered)} account name(s) reduce to the exact family "
                    f"signature {group_key!r} after legal and state suffix normalization.",
                ),
                requires_review=False,
            )
        )
    return tuple(sorted(proposals, key=lambda proposal: proposal.display_name.casefold()))


def load_customer_price_file_references(
    workbook_path: str | Path,
    *,
    sheet_name: str = "FILES",
) -> tuple[PriceFileReference, ...]:
    """Read current Excel price-file paths from the supplied path workbook."""

    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook does not contain sheet {sheet_name!r}")
        sheet = workbook[sheet_name]
        seen: set[str] = set()
        references: list[PriceFileReference] = []
        for (raw_value,) in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
            value = clean_text(str(raw_value)) if raw_value is not None else None
            if value is None:
                continue
            windows_path = PureWindowsPath(value)
            suffix = windows_path.suffix.casefold()
            if suffix not in {".xls", ".xlsx", ".xlsm"}:
                continue
            if any(part.casefold() == "old" for part in windows_path.parts):
                continue
            dedupe_key = str(windows_path).casefold()
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            references.append(
                PriceFileReference(
                    path=str(windows_path),
                    file_name=windows_path.name,
                    normalized_stem=normalize_name(windows_path.stem),
                )
            )
        return tuple(sorted(references, key=lambda item: item.path.casefold()))
    finally:
        workbook.close()


def _match_score(left: str | None, right: str | None) -> int:
    left_tokens = set(_significant_tokens(left))
    right_tokens = set(_significant_tokens(right))
    if not left_tokens or not right_tokens:
        return 0
    if left_tokens == right_tokens:
        return 100
    intersection = left_tokens & right_tokens
    if len(intersection) >= 2 and (
        intersection == left_tokens or intersection == right_tokens
    ):
        return 90
    jaccard = len(intersection) / len(left_tokens | right_tokens)
    left_text = " ".join(sorted(left_tokens))
    right_text = " ".join(sorted(right_tokens))
    sequence = SequenceMatcher(None, left_text, right_text).ratio()
    return round(max(jaccard, sequence * 0.85) * 100)


def _rank_alternatives(
    source_name: str,
    targets: Iterable[tuple[str, str]],
    *,
    minimum_score: int,
) -> tuple[MatchAlternative, ...]:
    alternatives = [
        MatchAlternative(target_key=key, target_name=name, score=_match_score(source_name, name))
        for key, name in targets
    ]
    return tuple(
        sorted(
            (candidate for candidate in alternatives if candidate.score >= minimum_score),
            key=lambda candidate: (-candidate.score, candidate.target_name.casefold()),
        )
    )


def _target_token_index(
    targets: Iterable[tuple[str, str]],
) -> tuple[dict[str, tuple[str, str]], dict[str, set[str]]]:
    """Index targets by meaningful tokens so full exports remain fast to assess."""

    by_key: dict[str, tuple[str, str]] = {}
    token_index: dict[str, set[str]] = defaultdict(set)
    for key, name in targets:
        by_key[key] = (key, name)
        for token in _significant_tokens(name):
            if len(token) >= 2:
                token_index[token].add(key)
    return by_key, token_index


def _candidate_targets(
    source_name: str,
    by_key: Mapping[str, tuple[str, str]],
    token_index: Mapping[str, set[str]],
) -> tuple[tuple[str, str], ...]:
    keys: set[str] = set()
    for token in _significant_tokens(source_name):
        if len(token) >= 2:
            keys.update(token_index.get(token, ()))
    return tuple(by_key[key] for key in sorted(keys))


def propose_customer_price_file_matches(
    groups: Iterable[CustomerGroupProposal],
    price_files: Iterable[PriceFileReference],
    *,
    minimum_score: int = 60,
    ambiguity_margin: int = 3,
) -> tuple[CustomerPriceFileProposal, ...]:
    targets = tuple((group.group_key, group.display_name) for group in groups)
    by_key, token_index = _target_token_index(targets)
    proposals: list[CustomerPriceFileProposal] = []
    for reference in price_files:
        candidates = _candidate_targets(reference.normalized_stem, by_key, token_index)
        alternatives = _rank_alternatives(
            reference.normalized_stem,
            candidates,
            minimum_score=minimum_score,
        )
        if not alternatives:
            proposals.append(
                CustomerPriceFileProposal(
                    file_path=reference.path,
                    file_name=reference.file_name,
                    customer_group_key=None,
                    customer_group_name=None,
                    score=0,
                    method="normalized_price_filename",
                    evidence=(
                        f"No customer group met the minimum score for price filename "
                        f"{reference.normalized_stem!r}.",
                    ),
                    alternatives=(),
                    requires_review=True,
                )
            )
            continue
        best = alternatives[0]
        ambiguous = len(alternatives) > 1 and alternatives[1].score >= best.score - ambiguity_margin
        selected = None if ambiguous else best
        proposals.append(
            CustomerPriceFileProposal(
                file_path=reference.path,
                file_name=reference.file_name,
                customer_group_key=selected.target_key if selected else None,
                customer_group_name=selected.target_name if selected else None,
                score=best.score,
                method="normalized_price_filename",
                evidence=(
                    f"Price filename stem {reference.normalized_stem!r} scored {best.score} "
                    f"against customer group {best.target_name!r}.",
                    "Multiple near-equal candidates require selection."
                    if ambiguous
                    else "The highest-scoring candidate is unique.",
                ),
                alternatives=alternatives[:5],
                requires_review=ambiguous or best.score < 100,
            )
        )
    return tuple(sorted(proposals, key=lambda proposal: proposal.file_path.casefold()))


def most_recent_purchase_by_item(
    purchases: Iterable[SourceRow],
) -> dict[str, PurchaseEvidence]:
    """Retain the most recent usable purchase row for each planning item."""

    latest: dict[str, PurchaseEvidence] = {}
    for row in purchases:
        item_number = clean_text(row.get("Item Number"))
        supplier_name = clean_text(row.get("Co./Last Name"))
        purchase_date = parse_date(row.get("Date"))
        if (
            item_number is None
            or supplier_name is None
            or purchase_date is None
            or is_control_item_number(item_number)
        ):
            continue
        evidence = PurchaseEvidence(
            item_number=item_number,
            supplier_name=supplier_name,
            purchase_number=clean_text(row.get("Purchase No.")),
            purchase_date=purchase_date,
            purchase_status=clean_text(row.get("Purchase Status")),
            price_text=clean_text(row.get("Price")),
            currency_code=clean_text(row.get("Currency Code")),
        )
        current = latest.get(item_number)
        if current is None or evidence.purchase_date > current.purchase_date:
            latest[item_number] = evidence
    return latest


def _supplier_targets(
    suppliers: Iterable[SupplierMasterCandidate],
) -> tuple[tuple[str, str], ...]:
    return tuple(
        (
            supplier.myob_record_id or f"name:{supplier.normalized_name}",
            supplier.display_name,
        )
        for supplier in suppliers
    )


def propose_item_supplier_matches(
    items: Iterable[ItemMasterCandidate],
    suppliers: Iterable[SupplierMasterCandidate],
    purchases: Iterable[SourceRow],
    *,
    minimum_score: int = 65,
    ambiguity_margin: int = 3,
) -> tuple[ItemSupplierProposal, ...]:
    """Use MYOB primary supplier first, then the most recent purchase supplier."""

    supplier_targets = _supplier_targets(suppliers)
    supplier_by_key, supplier_token_index = _target_token_index(supplier_targets)
    recent_by_item = most_recent_purchase_by_item(purchases)
    proposals: list[ItemSupplierProposal] = []
    for item in items:
        if item.excluded_from_item_view:
            continue
        last_purchase = recent_by_item.get(item.item_number)
        source_supplier_name = clean_text(item.primary_supplier_name)
        method = "myob_primary_supplier"
        if source_supplier_name is None and last_purchase is not None:
            source_supplier_name = last_purchase.supplier_name
            method = "most_recent_purchase_supplier"
        if source_supplier_name is None:
            continue

        candidates = _candidate_targets(
            source_supplier_name,
            supplier_by_key,
            supplier_token_index,
        )
        alternatives = _rank_alternatives(
            source_supplier_name,
            candidates,
            minimum_score=minimum_score,
        )
        if not alternatives:
            proposals.append(
                ItemSupplierProposal(
                    item_number=item.item_number,
                    source_supplier_name=source_supplier_name,
                    supplier_record_id=None,
                    supplier_name=None,
                    score=0,
                    method=method,
                    evidence=(
                        f"No supplier card met the minimum score for {source_supplier_name!r}.",
                    ),
                    alternatives=(),
                    last_purchase=last_purchase,
                    requires_review=True,
                )
            )
            continue

        best = alternatives[0]
        ambiguous = len(alternatives) > 1 and alternatives[1].score >= best.score - ambiguity_margin
        selected = None if ambiguous else best
        evidence = [
            f"{method.replace('_', ' ').title()} supplied name {source_supplier_name!r}.",
            f"Best supplier-card score is {best.score} for {best.target_name!r}.",
        ]
        if last_purchase is not None:
            evidence.append(
                f"Latest purchase evidence is {last_purchase.purchase_number or 'unknown'} "
                f"dated {last_purchase.purchase_date.isoformat()} with status "
                f"{last_purchase.purchase_status or 'unknown'}."
            )
        if ambiguous:
            evidence.append("Multiple near-equal supplier cards require selection.")
        proposals.append(
            ItemSupplierProposal(
                item_number=item.item_number,
                source_supplier_name=source_supplier_name,
                supplier_record_id=selected.target_key if selected else None,
                supplier_name=selected.target_name if selected else None,
                score=best.score,
                method=method,
                evidence=tuple(evidence),
                alternatives=alternatives[:5],
                last_purchase=last_purchase,
                requires_review=ambiguous or best.score < 100,
            )
        )
    return tuple(sorted(proposals, key=lambda proposal: proposal.item_number.casefold()))


def purchase_price_value(evidence: PurchaseEvidence | None) -> Decimal | None:
    """Return a parsed price without making currency assumptions."""

    return parse_decimal(evidence.price_text) if evidence is not None else None
